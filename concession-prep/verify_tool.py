# -*- coding: utf-8 -*-
"""
生成したExcel(v5)の計算結果を、CSVからPythonで独立集計した期待値と突合する。
使い方:
  python verify_tool.py <xlsx> --template
  python verify_tool.py <xlsx> --csv-a A.csv --csv-b B.csv --select A \
      --att-a 9000,13000,12000 --att-b ... --peak 1200 --mult 1.2 \
      --mso1 金曜1週目.csv --band 2         # 商品別の波(帯②)適用時の作る数も突合
"""
import argparse
import math
import sys

from openpyxl import load_workbook

from build_calib import CALIB_SHEETS, MSO_MAX, read_mso_rows
from build_tool import (DEFAULT_PRODUCTS, EXCLUDE_CATS, JUDGE_FEW, JUDGE_NODATA,
                        JUDGE_NONE, JUDGE_USE, WAVE_ROW0, WAVE_SHEET, read_csv_rows)

ap = argparse.ArgumentParser()
ap.add_argument("xlsx")
ap.add_argument("--template", action="store_true")
ap.add_argument("--csv-a")
ap.add_argument("--csv-b")
ap.add_argument("--select", choices=["A", "B", "AVG"], default="A")
ap.add_argument("--att-a")
ap.add_argument("--att-b")
ap.add_argument("--peak", type=int)
ap.add_argument("--mult", type=float, default=1.0, help="時間帯係数")
for k in range(1, 5):
    ap.add_argument(f"--mso{k}", help="係数貼付①〜④に入っているMSO商品CSV")
ap.add_argument("--band", type=int, default=0,
                help="選択中プリセットの帯番号1〜5(商品別の波の適用検証用。0=適用なし)")
ap.add_argument("--wave-thr", type=int, default=30, help="商品別係数を使う最低個数")
ap.add_argument("--close", default="26:00",
                help="係数算出H4と同じ閉店時刻(22:00=同日/26:00=翌2:00。既定26:00)")
ap.add_argument("--open", dest="open_", default="08:00",
                help="係数算出C4と同じ開店時刻(既定08:00)")
a = ap.parse_args()

wb = load_workbook(a.xlsx, data_only=True)
errors = []

# 開店・閉店時刻→帯①/⑤の窓と帯長(Excel側のC4/N4/O4と同じルール)
_h, _m = a.open_.split(":")
OPEN_H = int(_h) + int(_m) / 60
_h, _m = a.close.split(":")
_close_h = int(_h) + int(_m) / 60
if _close_h >= 24:
    CAP5, ND5, DUR5 = 1.0, (_close_h - 24) / 24, 3 + (_close_h - 24)
elif _close_h >= 21:
    CAP5, ND5, DUR5 = _close_h / 24, 0.0, _close_h - 21
else:                       # 翌側表記(2:00等)
    CAP5, ND5, DUR5 = 1.0, _close_h / 24, 3 + _close_h
DURS = [11 - OPEN_H, 4, 3, 3, DUR5]


def check(label, got, want, tol=0):
    ok = (got == want) if tol == 0 else (
        got is not None and want is not None and abs(got - want) <= tol)
    if not ok:
        errors.append(f"NG {label}: got={got!r} want={want!r}")


def sales_by_name(csv_path):
    total = {}
    for row in read_csv_rows(csv_path):
        name, qty = row[13], row[26]
        if name is not None:
            total[name] = total.get(name, 0) + (qty or 0)
    return total


def mso_rows_filtered(path):
    """Excel側AF列と同じ対象行(先頭行の日付・提供済・販売・セット親以外・数量>0)。
    時刻はAE式と同じくMOD/TIMEVALUE相当で24時間にラップし、解釈不能な行は除外"""
    rows = read_mso_rows(path)[:MSO_MAX]
    target = rows[0][10] if rows else None
    out = []
    for r in rows:
        if (r[10] != target or r[18] != "提供済" or r[19] != "販売"
                or r[27] == "セット親"):
            continue
        q = r[25] or 0
        if not isinstance(q, (int, float)) or q <= 0:
            continue
        t = r[12]
        if isinstance(t, (int, float)):
            tv = t % 1.0
        elif isinstance(t, str):
            try:
                parts = [int(x) for x in t.split(":")]
            except ValueError:
                continue
            if len(parts) == 2:
                parts.append(0)
            if len(parts) != 3:
                continue
            tv = ((parts[0] * 3600 + parts[1] * 60 + parts[2]) / 86400) % 1.0
        else:
            continue
        out.append((tv, q, r[23]))
    return out


def band_of(tv):
    windows = [(OPEN_H / 24, 11 / 24), (11 / 24, 15 / 24), (15 / 24, 18 / 24), (18 / 24, 21 / 24)]
    for bi, w in enumerate(windows):
        if w[0] <= tv < w[1]:
            return bi
    if (21 / 24 <= tv < CAP5) or tv < ND5:
        return 4
    return None


mso_paths = [a.mso1, a.mso2, a.mso3, a.mso4]
week_counts = []          # 週ごとの帯別合計
prod_counts = {}          # 商品名 → 帯別個数(全有効週合計)
for p in mso_paths:
    wcs = [0] * 5
    if p:
        for tv, q, name in mso_rows_filtered(p):
            b = band_of(tv)
            if b is None:
                continue
            wcs[b] += q
            if name:
                prod_counts.setdefault(name, [0] * 5)[b] += q
    week_counts.append(wcs)
n_weeks = sum(1 for wcs in week_counts if sum(wcs) > 0)


def check_r005(label, got, raw):
    """0.05刻みセルの検証。刻みのちょうど中間は浮動小数の微差で丸め方向が
    環境依存になるため、「0.05の倍数」かつ「生係数から半ステップ以内」を確認する"""
    ok = (isinstance(got, (int, float))
          and abs(got / 0.05 - round(got / 0.05)) < 1e-6
          and abs(got - raw) <= 0.025 + 1e-6)
    if not ok:
        errors.append(f"NG {label}: got={got!r} raw={raw!r}")


def wave_coef(name, band):
    """商品別係数(丸め前の生値)。使えないとき(帯個数0や、0.05刻みの丸めで
    0.00になる係数のフォールバック含む)はNone"""
    if n_weeks == 0 or band < 1 or band > 5:
        return None
    cs = prod_counts.get(name)
    total = sum(cs) if cs else 0
    if total == 0 or total < a.wave_thr or cs[band - 1] == 0:
        return None
    raw = (cs[band - 1] / total) * (sum(DURS) / DURS[band - 1])
    if raw / 0.05 < 0.5 - 1e-9:     # Excel側はIF(ROUND(...)=0,"—",...)
        return None
    return raw


pd = wb["期間データ"]
m = wb["準備数計算"]
pr = wb["印刷用"]
wv = wb[WAVE_SHEET]

if a.template:
    for i in range(len(DEFAULT_PRODUCTS)):
        r = 14 + i
        check(f"期間データ!B{r}", pd[f"B{r}"].value, DEFAULT_PRODUCTS[i])
        check(f"期間データ!C{r}(A販売=0)", pd[f"C{r}"].value, 0)
        check(f"期間データ!D{r}(B販売=0)", pd[f"D{r}"].value, 0)
    check("期間データ!G4(A未貼付表示)", "貼り付けてください" in (pd["G4"].value or ""), True)
    check("準備数計算!E11(要確認)", m["E11"].value, "要確認")
    check("準備数計算!F11(空欄)", m["F11"].value in (None, ""), True)
    check("準備数計算!H11(商品係数=—)", m["H11"].value, "—")
    warn = m["B9"].value or ""
    for kw in ("未貼付", "そろっていません", "ピーク動員数"):
        check(f"B9警告[{kw}]", kw in warn, True)
else:
    sa = sales_by_name(a.csv_a)
    sb = sales_by_name(a.csv_b)
    att_a = [int(x) for x in a.att_a.split(",")]
    att_b = [int(x) for x in a.att_b.split(",")]
    if a.select == "AVG":
        att_sel = sum(att_a) + sum(att_b)
        sel_sales = {k: sa.get(k, 0) + sb.get(k, 0) for k in set(sa) | set(sb)}
    else:
        att_sel = sum(att_a) if a.select == "A" else sum(att_b)
        sel_sales = sa if a.select == "A" else sb
    # 参考列(比較期間): A選択時は期間B、それ以外は期間A
    att_oth = sum(att_b) if a.select == "A" else sum(att_a)
    oth_sales = sb if a.select == "A" else sa

    check("期間データ!J6(A動員計)", pd["J6"].value, sum(att_a))
    check("期間データ!J10(B動員計)", pd["J10"].value, sum(att_b))
    check("期間データ!G4(✔3日間)", "✔ 3日間" in (pd["G4"].value or ""), True)
    check("期間データ!B11(✔7日間)", "✔ 7日間" in (pd["B11"].value or ""), True)
    import datetime as _dt
    check("期間データ!C4(自動日付)", pd["C4"].value, _dt.datetime(2026, 8, 21))
    check("期間データ!E4(自動日付)", pd["E4"].value, _dt.datetime(2026, 8, 23))
    check("期間データ!C8(自動日付)", pd["C8"].value, _dt.datetime(2026, 8, 14))
    check("期間データ!I8(自動日付)", pd["I8"].value, _dt.datetime(2026, 8, 20))
    check("期間データ!C37(除外リスト照合✔)", pd["C37"].value, "✔")

    for i, name in enumerate(DEFAULT_PRODUCTS):
        r = 14 + i
        check(f"期間データ!C{r}({name[:6]})", pd[f"C{r}"].value, sa.get(name, 0))
        check(f"期間データ!D{r}", pd[f"D{r}"].value, sb.get(name, 0))

    check("準備数計算!M4", m["M4"].value, {"A": 1, "B": 2, "AVG": 3}[a.select])
    check("準備数計算!M5", m["M5"].value, att_sel)
    check("準備数計算!M6", m["M6"].value, att_oth)
    check("準備数計算!M7(時間帯係数)", m["M7"].value, a.mult, tol=1e-9)
    wave_active = a.band >= 1 and a.band <= 5 and n_weeks > 0
    check("準備数計算!M9(商品別波の適用)", m["M9"].value, 1 if wave_active else 0)
    for i, name in enumerate(DEFAULT_PRODUCTS):
        r = 11 + i
        s = sel_sales.get(name, 0)
        rate = s / att_sel
        cp = wave_coef(name, a.band) if wave_active else None
        hv = m[f"H{r}"].value
        if cp is not None:
            # H列(0.05刻み)を生係数で検証し、作る数はExcelが実際に使った係数で突合
            check_r005(f"準備数計算!H{r}(商品係数)", hv, cp)
            eff = hv if isinstance(hv, (int, float)) else 1.0
        else:
            check(f"準備数計算!H{r}(商品係数=—)", hv, "—")
            # 波の適用中(M9=1)は時間帯係数を使わない排他設計: 係数の出ない商品は×1.0
            eff = 1.0 if wave_active else a.mult
        check(f"準備数計算!D{r}", m[f"D{r}"].value, s)
        check(f"準備数計算!E{r}", m[f"E{r}"].value, rate, tol=1e-9)
        check(f"準備数計算!F{r}", m[f"F{r}"].value, math.ceil(a.peak * rate * eff), tol=1)
        check(f"準備数計算!G{r}", m[f"G{r}"].value, oth_sales.get(name, 0) / att_oth, tol=1e-9)
        check(f"印刷用!D{8 + i}", pr[f"D{8 + i}"].value, math.ceil(a.peak * rate * eff), tol=1)
    warn = m["B9"].value
    if warn not in (None, ""):
        errors.append(f"NG B9警告が出ている: {warn!r}")

    # 商品リスト(プルダウン)の中身: 除外小分類を除いた先出順
    def uniq_names(path):
        return list(dict.fromkeys(
            r[13] for r in read_csv_rows(path)
            if r[13] and r[7] not in EXCLUDE_CATS))
    uniq_a = uniq_names(a.csv_a)
    uniq_b = uniq_names(a.csv_b)
    check("商品リスト先頭", pd["Q5"].value, uniq_a[0])
    check("商品リストA件数目", pd[f"Q{4 + len(uniq_a)}"].value, uniq_a[-1])
    check("商品リストB先頭", pd[f"Q{5 + len(uniq_a)}"].value, uniq_b[0])
    check("商品リスト末尾", pd[f"Q{4 + len(uniq_a) + len(uniq_b)}"].value, uniq_b[-1])
    check("商品リスト末尾+1は空", pd[f"Q{5 + len(uniq_a) + len(uniq_b)}"].value in (None, ""), True)
    # 除外カテゴリの商品がリストに無いこと
    excluded_names = {r[13] for r in read_csv_rows(a.csv_a) if r[13] and r[7] in EXCLUDE_CATS}
    listed = {pd.cell(row=rr, column=17).value for rr in range(5, 5 + len(uniq_a) + len(uniq_b))}
    leak = excluded_names & listed
    if leak:
        errors.append(f"NG 除外カテゴリの商品がリストに混入: {sorted(leak)[:5]}")

# 空き枠
for i in range(len(DEFAULT_PRODUCTS), 20):
    for col in "CDEFGH":
        v = m[f"{col}{11 + i}"].value
        if v not in (None, ""):
            errors.append(f"NG 準備数計算 空き枠 {col}{11 + i}: {v!r}")

# ------------------------------------------------ 係数算出(時間帯係数の較正) --
ks = wb["係数算出"]
for s in CALIB_SHEETS:
    check(f"{s}!AD3(正常フラグ)", wb[s]["AD3"].value, 1)

if any(mso_paths):
    for wi, wcs in enumerate(week_counts):
        col = "DEFG"[wi]
        for bi in range(5):
            check(f"係数算出!{col}{7 + bi}({wi + 1}週目)", ks[f"{col}{7 + bi}"].value, wcs[bi])
        check(f"係数算出!{col}12(週合計)", ks[f"{col}12"].value, sum(wcs))
    check("係数算出!H13(週数)", ks["H13"].value, n_weeks)
    if n_weeks:
        avgs = [sum(wcs[bi] for wcs in week_counts) / n_weeks for bi in range(5)]
        day_pace = sum(avgs) / sum(DURS)
        for bi in range(5):
            coef = (avgs[bi] / DURS[bi]) / day_pace
            check(f"係数算出!K{7 + bi}(係数候補)", ks[f"K{7 + bi}"].value, coef, tol=1e-6)
            lv = ks[f"L{7 + bi}"].value
            check_r005(f"係数算出!L{7 + bi}(転記用)", lv, coef)
            check(f"準備数計算!K{4 + bi}(実測候補の連動)", m[f"K{4 + bi}"].value, lv, tol=1e-9)
        check("係数算出!B23(警告なし)", ks["B23"].value in (None, ""), True)
        check("係数算出!N5(区切り外=0)", ks["N5"].value, 0)
        check("係数算出!O5(同日疑い週=0)", ks["O5"].value, 0)
    for wi, p in enumerate(mso_paths):
        st = wb[CALIB_SHEETS[wi]]["A3"].value or ""
        if p:
            check(f"{CALIB_SHEETS[wi]}!A3(貼付済表示)", "貼付" in st and "対象個数" in st, True)
        else:
            check(f"{CALIB_SHEETS[wi]}!A3(未貼付表示)", "未貼付" in st, True)
    # 商品別の波: 帯別個数(非表示P〜T)・合計・構成比・係数・判定
    for i, name in enumerate(DEFAULT_PRODUCTS):
        wr = WAVE_ROW0 + i
        cs = prod_counts.get(name, [0] * 5)
        total = sum(cs)
        for bi in range(5):
            pcol = "PQRST"[bi]
            check(f"{WAVE_SHEET}!{pcol}{wr}({name[:5]})", wv[f"{pcol}{wr}"].value, cs[bi])
            ccol = "CDEFG"[bi]
            if total == 0:
                check(f"{WAVE_SHEET}!{ccol}{wr}(構成比—)", wv[f"{ccol}{wr}"].value, "—")
            else:
                check(f"{WAVE_SHEET}!{ccol}{wr}(構成比)", wv[f"{ccol}{wr}"].value, cs[bi] / total, tol=1e-9)
            kcol = "IJKLM"[bi]
            cexp = wave_coef(name, bi + 1)
            if cexp is None:
                check(f"{WAVE_SHEET}!{kcol}{wr}(係数—)", wv[f"{kcol}{wr}"].value, "—")
            else:
                check_r005(f"{WAVE_SHEET}!{kcol}{wr}(係数)", wv[f"{kcol}{wr}"].value, cexp)
        check(f"{WAVE_SHEET}!H{wr}(合計)", wv[f"H{wr}"].value, total)
        want_j = (JUDGE_NODATA if n_weeks == 0 else
                  JUDGE_NONE if total == 0 else
                  JUDGE_FEW if total < a.wave_thr else JUDGE_USE)
        check(f"{WAVE_SHEET}!N{wr}(判定)", wv[f"N{wr}"].value, want_j)
else:
    check("係数算出!H13(週数=0)", ks["H13"].value, 0)
    check("係数算出!N5(区切り外=0)", ks["N5"].value, 0)
    check("係数算出!O5(同日疑い週=0)", ks["O5"].value, 0)
    check("係数算出!L7(未算出=—)", ks["L7"].value, "—")
    check("係数算出!B23(未貼付警告)", "データが貼られていません" in (ks["B23"].value or ""), True)
    check("準備数計算!K4(実測候補=—)", m["K4"].value, "—")
    check("準備数計算!M9(商品別波=不使用)", m["M9"].value, 0)
    for s in CALIB_SHEETS:
        check(f"{s}!A3(未貼付)", wb[s]["A3"].value, "（未貼付）")
    if (wv["B7"].value or "") != "":
        check(f"{WAVE_SHEET}!N7(MSO未貼付)", wv["N7"].value, JUDGE_NODATA)
        check(f"{WAVE_SHEET}!I7(係数—)", wv["I7"].value, "—")

if errors:
    print(f"FAILED ({len(errors)}):")
    for e in errors[:40]:
        print(" ", e)
    sys.exit(1)
print("ALL OK", "(template)" if a.template else f"(select={a.select}, 動員={att_sel}, band={a.band})")
