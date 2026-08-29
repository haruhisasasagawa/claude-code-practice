# -*- coding: utf-8 -*-
"""
TOHOシネマズ新宿 コンセッション事前準備数ツール（v5.2）生成スクリプト

TOHOの営業週(金曜開始)に合わせ、本社集計「売上・在庫・原価」CSV(34列/cp932)を
期間A(直近金土日)・期間B(前週 金〜木)の2本貼り付けて使う構成。

シート構成:
  使い方       … 3ステップの利用ガイド・凡例・注意点
  準備数計算   … 参照期間(A/B)の購買率 × ピーク動員数 × 係数(時間帯/商品別の波) →「作る数」
  印刷用       … A4縦1枚の仕込み指示書(自動連動・チェック欄付き)
  期間データ   … 期間A/Bの日付・動員数を入力。商品別販売数はCSVから自動集計
  CSV貼付A/B   … 集計CSVをそのまま貼るだけの貼り付けシート(数式なし)
  係数算出     … 金曜4週分のMSO商品CSVから時間帯係数の実測候補を算出(月1回・任意)
  商品別の波   … 登録商品ごとの時間帯パターン(構成比・商品別係数)。作る数への適用は
                 準備数計算の「④ 商品別の波」(D7)で切替。係数は商品ごとに置き換えで、
                 商品係数が出ない商品は全体の時間帯係数で計算
  係数貼付①〜④ … MSO商品CSV(注文明細)を1週1日分ずつ貼る較正用シート(build_calib.py)

使い方:
  python build_tool.py                            # 未入力テンプレートを生成
  python build_tool.py --out サンプル.xlsx \
      --csv-a A期間.csv --csv-b B期間.csv \
      --select A --att-a 9000,13000,12000 --att-b 8500,... --peak 1200 \
      --preset "② 昼ピーク（11〜15時）" \
      --mso1 0717MSO.csv                          # 係数貼付①にMSO明細を投入
"""
import argparse
import csv
import datetime as dt

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

# ---------------------------------------------------------------- palette ----
FONT_NAME = "Meiryo UI"          # 全シート共通フォント。無い環境では自動代替

NAVY = "2E3A59"
INK = "333B4A"
GRAY = "8A93A3"
CORAL = "E8604C"
TEAL = "1F9E92"
AMBER = "E9A13B"
AMBER2 = "C77E23"
GREEN = "2E9E5B"
RED = "D14343"

F_INPUT = "FFF4D6"
B_INPUT = "F0D48A"
F_AUTO = "F2F4F7"
F_ZEBRA = "F8FAFC"
F_BASE = "FFF1EE"
CHIP_NAVY = "E9EDF5"
CHIP_AMBER = "FDF0DA"
CHIP_CORAL = "FCE7E2"
CHIP_TEAL = "DFF2F0"
LINE = "DFE3EA"

thin = Side(style="thin", color=LINE)
hair = Side(style="hair", color=LINE)
in_side = Side(style="thin", color=B_INPUT)
coral_side = Side(style="medium", color=CORAL)
BORDER_LIGHT = Border(left=thin, right=thin, top=thin, bottom=thin)
BORDER_HAIR = Border(left=hair, right=hair, top=hair, bottom=hair)
BORDER_INPUT = Border(left=in_side, right=in_side, top=in_side, bottom=in_side)
UNLOCKED = Protection(locked=False)


def fnt(size=10.5, bold=False, color=INK):
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def fill(color):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def coord(ref):
    col = "".join(ch for ch in ref if ch.isalpha())
    row = int("".join(ch for ch in ref if ch.isdigit()))
    return column_index_from_string(col), row


def style_range(ws, ref, font=None, fl=None, alignment=None, border=None, num=None):
    start, end = ref.split(":") if ":" in ref else (ref, ref)
    sc, sr = coord(start)
    ec, er = coord(end)
    for r in range(sr, er + 1):
        for c in range(sc, ec + 1):
            cell = ws.cell(row=r, column=c)
            if font:
                cell.font = font
            if fl:
                cell.fill = fl
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            if num:
                cell.number_format = num


def title_band(ws, ref, text):
    ws.merge_cells(ref)
    style_range(ws, ref, font=fnt(14, True, "FFFFFF"), fl=fill(NAVY),
                alignment=align("left", "center"))
    ws[ref.split(":")[0]] = text


def chip(ws, ref, text, chip_fill, color=INK, size=10, bold=True, h="left"):
    if ":" in ref:
        ws.merge_cells(ref)
    style_range(ws, ref, font=fnt(size, bold, color), fl=fill(chip_fill),
                alignment=align(h, "center"))
    ws[ref.split(":")[0]] = text


def note(ws, ref, text, size=9, color=GRAY, wrap=False, h="left"):
    if ":" in ref:
        ws.merge_cells(ref)
    style_range(ws, ref, font=fnt(size, False, color),
                alignment=align(h, "center", wrap))
    ws[ref.split(":")[0]] = text


def disp_w(text):
    """表示幅(全角=1、半角=0.5)。len()だと半角混じりの行数を過大に見積もり、
    1行で収まる項目に2行分の高さが付いて縦のリズムが崩れる。"""
    return sum(0.5 if ord(ch) < 256 else 1 for ch in text)


# ------------------------------------------------------------- constants -----
CSV_HEADERS = ["タイトル", "劇場コード", "劇場名", "対象期間開始", "対象期間終了",
               "大分類コード", "小分類コード", "小分類名", "作品コード", "作品名",
               "支払先コード", "支払先名", "商品コード", "商品名", "買取区分", "買取区分名",
               "前日残数", "仕入数", "移動入庫数", "返品数", "廃棄数", "移動出庫数",
               "差異数", "資産廃棄数", "減算数", "残数", "売上数", "売上金額（税込）",
               "軽減売上数", "軽減売上金額（税込）", "売上金額（税抜）", "平均単価（税抜）",
               "売上原価（税抜）", "仕入先商品コード"]
NCOL = len(CSV_HEADERS)          # 34列: N=商品名(14), AA=売上数(27), D/E=対象期間(4/5)
CSV_MAX = 1000                   # 貼付データ最大行数(5〜1004行目)
CSV_END = 4 + CSV_MAX
LIST_MAX = CSV_MAX               # 期間ごとの商品リスト最大件数(全行ユニークでも切れない)
N_SLOTS = 20                     # 商品枠
ROW_P0 = 14                      # 期間データ: 商品1行目(14〜33)
ROW_M0 = 11                      # 準備数計算: 商品1行目(11〜30)

# 事前準備(調理・仕込み)対象になりやすい実売上位の商品(実CSVの商品名と完全一致)
DEFAULT_PRODUCTS = [
    "ハーフ＆ハーフ（塩／キャラメル）",
    "ポップコーン　キャラメルＭ",
    "ポップコーン　塩Ｍ",
    "ポップコーン　キャラメルＳ",
    "ポップコーン　塩Ｓ",
    "ポップコーン　キャラメルＬ",
    "ポップコーン　塩Ｌ",
    "北海道濃厚バターしょうゆ味",
    "トリュフソルト＆バター味",
    "シネマイク　ハッピーターン味",
    "パーティーポップ　キャラメル",
    "スパイシー！ポップチキン",
    "スナックじゃがトリュフソルトバタ",
    "スナックじゃが　シチリアハーブ",
    "ケチャップ＆マスタード",
    "４種のチーズ",
    "プレミアム　ハラペーニョ",
    "チュリトス　シナモンシュガー",
    "チュリトス　チョコクリーム",
]

# 時間帯は5段階(朝一→昼ピーク→夕方→夜ピーク→レイト)+基準の平常。名前・係数・時刻の目安は編集可
PRESETS = [("① 朝一（〜11時）", 0.8),
           ("② 昼ピーク（11〜15時）", 1.2),
           ("③ 夕方（15〜18時）", 1.1),
           ("④ 夜ピーク（18〜21時）", 1.3),
           ("⑤ レイト（21時〜）", 0.9),
           ("平常（基準）", 1.0)]
PRESET_SLOTS = 7                 # プリセット表の枠数(I4:J10。6件+空き1枠)
PRESET_END = 3 + PRESET_SLOTS

# プルダウン検索から既定で除外する小分類(実CSVのH列の表記に完全一致・シート上で編集可)
EXCLUDE_CATS = ["コールド", "コーヒー", "アルコール", "その他ドリンク", "ホット",
                "ドリンク調味料", "フード調味料", "ＳＥＴ作品コンボ", "引換券", "コンセ包材"]
EXC_SLOTS = 12                   # 除外リストの枠数(期間データ!B37:B48)
EXC_TOP = 37
EXC_END = EXC_TOP + EXC_SLOTS - 1

SEL_A = "期間A（直近金土日）"
SEL_B = "期間B（金〜木）"
SEL_AVG = "期間平均（A+B）"

# 商品別の波(商品ごとの時間帯係数)シートとの連携
WAVE_SHEET = "商品別の波"
WAVE_ROW0 = 7                    # 商品1行目(7〜26。準備数計算11〜30と同順)
WAVE_ON = "使う"
WAVE_OFF = "使わない"
JUDGE_USE = "✔ 商品別係数を使用"
JUDGE_FEW = "⚠ 少データ → 全体係数"
JUDGE_NONE = "該当なし → 全体係数"
JUDGE_NODATA = "（MSO未貼付）"

WEEKDAY_JA = '"日","月","火","水","木","金","土"'


def parse_clock(s):
    """"22:00"/"26:00"/"2:00" → 時刻シリアル値(時/24)。24時間超え表記も可"""
    h, m = s.split(":")
    return (int(h) + int(m) / 60) / 24


def mk_comment(text, width=300):
    """内容量に応じた吹き出しサイズのメモ。文字は仕上げ(force_font)で
    Meiryo UI 9ptに統一されるため、その字面を前提に高さを見積もる"""
    import math
    lines = sum(max(1, math.ceil(len(seg) / 24)) for seg in text.split("\n"))
    return Comment(text, "準備数ツール", height=min(380, lines * 15 + 16), width=width)


def parse_ymd(x):
    """yyyymmdd数値/文字列・日付型セルのどれでも日付シリアルに解釈する式。
    2000〜2100年の範囲外(0・空欄・壊れた値)は-1を返す。"""
    inner = (f'IF(VALUE({x})<19000101,INT(VALUE({x})),'
             f'DATEVALUE(TEXT(VALUE({x}),"0000-00-00")))')
    return f'IFERROR(IF(AND({inner}>=36526,{inner}<=73415),{inner},-1),-1)'


PASTE_SHEETS = ("CSV貼付A", "CSV貼付B", "係数貼付①", "係数貼付②", "係数貼付③", "係数貼付④")


def force_font(path, name=None):
    """ワークブック内の全フォント名を指定フォントへ書き戻す。
    LibreOfficeでの再計算保存時にセルのフォント名が環境の代替フォント
    (WenQuanYi Zen Hei等)へ置換されるため、styles.xml(セル・条件付き書式)と
    テーマのフォント名をzipレベルで一括修正する。再計算後の仕上げに呼ぶこと。"""
    import os
    import re
    import zipfile

    name = name or FONT_NAME
    tmp = path + ".tmpf"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = re.sub(rb'<name val="[^"]*"',
                              f'<name val="{name}"'.encode(), data)
            elif item.filename.startswith("xl/theme/"):
                data = re.sub(rb'typeface="[^"]+"',
                              f'typeface="{name}"'.encode(), data)
            elif re.fullmatch(r"xl/comments\d+\.xml", item.filename):
                # メモ(コメント)本文もフォント名が代替に置換されるため書き戻し、
                # 文字サイズを9ptに統一する
                data = re.sub(rb'<rFont val="[^"]*"',
                              f'<rFont val="{name}"'.encode(), data)
                data = re.sub(rb'<sz val="[^"]*"', b'<sz val="9"', data)
            zout.writestr(item, data)
    os.replace(tmp, path)


def show_paste_comments(path, sheets=PASTE_SHEETS):
    """指定シートのコメント(A4の貼り付け案内)を常時表示にする。
    openpyxl/LibreOfficeはコメントを非表示で書き出すため、xlsx内のVMLを直接
    書き換える。生成・再計算・後編集がすべて終わった最後に呼ぶこと。"""
    import os
    import re
    import zipfile

    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        wbxml = z.read("xl/workbook.xml").decode("utf-8")
        rels = z.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        rid_target = dict(re.findall(r'Id="([^"]+)"[^>]*?Target="([^"]+)"', rels))
        sheet_files = {}
        for tag in re.findall(r"<sheet [^>]*/>", wbxml):
            nm = re.search(r'name="([^"]+)"', tag)
            rid = re.search(r'r:id="([^"]+)"', tag)
            if nm and rid and rid.group(1) in rid_target:
                tgt = rid_target[rid.group(1)]
                sheet_files[nm.group(1)] = tgt if tgt.startswith("xl/") else "xl/" + tgt.lstrip("/")
        vml_files = set()
        for s in sheets:
            sf = sheet_files.get(s)
            if not sf:
                continue
            d, f = sf.rsplit("/", 1)
            rel = f"{d}/_rels/{f}.rels"
            if rel not in names:
                continue
            for t in re.findall(r'Target="([^"]+\.vml)"', z.read(rel).decode("utf-8")):
                vml_files.add(os.path.normpath(os.path.join(d, t)).replace("\\", "/"))
    if not vml_files:
        return 0
    patched = 0
    tmp = path + ".tmp"
    with zipfile.ZipFile(path) as zin, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename in vml_files and b"visibility:hidden" in data:
                data = data.replace(b"visibility:hidden", b"visibility:visible")
                # 常時表示メモがタイトル・貼付先A4を覆わないよう、位置を
                # A4の右下(データ領域の上)へ移し、3行分の横長サイズに整える
                data = re.sub(rb"<x:Anchor>[^<]*</x:Anchor>",
                              b"<x:Anchor>3, 12, 4, 6, 8, 40, 8, 8</x:Anchor>", data)
                data = re.sub(rb"margin-left:[^;'\"]*", b"margin-left:130pt", data)
                data = re.sub(rb"margin-top:[^;'\"]*", b"margin-top:104pt", data)
                data = re.sub(rb"width:[0-9.]+p[tx]", b"width:330pt", data)
                data = re.sub(rb"height:[0-9.]+p[tx]", b"height:78pt", data)
                patched += 1
            zout.writestr(item, data)
    os.replace(tmp, path)
    return patched


def read_csv_rows(path):
    """cp932/utf-8のCSVを読み、値を数値化して返す(ヘッダー行は除く)"""
    raw = open(path, "rb").read()
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    rows = list(csv.reader(text.splitlines()))
    out = []
    for r in rows[1:]:
        conv = []
        for v in r:
            v = v.strip()
            if v == "":
                conv.append(None)
                continue
            try:
                f = float(v)
                conv.append(int(f) if f == int(f) else f)
            except ValueError:
                conv.append(v)
        out.append(conv)
    return out


# ---------------------------------------------------------------- build ------
def build(out_path, csv_a=None, csv_b=None, select="A",
          att_a=None, att_b=None, peak=None, preset="平常（基準）",
          products=DEFAULT_PRODUCTS, mso_csvs=(None, None, None, None),
          close="22:00", open_="08:00"):
    wb = Workbook()

    # ============================================================ 使い方 =====
    ws = wb.active
    ws.title = "使い方"
    ws.sheet_properties.tabColor = NAVY
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    # 縦長の文章シートのため縦向き(横向きだと縮小されて右半分が空白になる)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = 9
    for c, w in {"A": 2.5, "B": 6, "C": 13, "D": 13, "E": 13, "F": 13, "G": 13,
                 "H": 13, "I": 13, "J": 13}.items():
        ws.column_dimensions[c].width = w

    ws.row_dimensions[1].height = 38
    title_band(ws, "A1:J1", "　🍿 コンセッション 事前準備数ツール")
    ws.row_dimensions[2].height = 20
    note(ws, "B2:J2", "TOHOシネマズ新宿｜期間A(直近金土日)・期間B(前週 金〜木)の購買率から、ピーク前の仕込み数を自動計算", 9.5)

    ws.row_dimensions[4].height = 22
    chip(ws, "B4:D4", "  つかいかた（3ステップ）", CHIP_NAVY, NAVY)
    steps = [
        ("①", AMBER, "「CSV貼付A」「CSV貼付B」に 売上・在庫・原価CSV を貼り付け",
         "CSVを全選択コピーし、オレンジのA4セルを選択→『値の貼り付け』（ヘッダー行ごとでOK）。"
         "日付は自動で入ります。｜担当：社員"),
        ("②", CORAL, "「期間データ」シートに 動員数 を入力",
         "期間A＝3日分、期間B＝7日分の動員数。日付・販売数は自動表示。CSVの行数・日数チェック（✔／⚠）も確認。｜担当：社員"),
        ("③", TEAL, "「準備数計算」で 参照期間・ピーク動員数・時間帯 を選ぶ",
         "ピーク動員数＝これから準備する回（例：1時間後のピーク）の合計動員数。「印刷用」をA4で刷って現場へ。｜担当：社員（確認：スタッフ）"),
    ]
    r = 6
    for mark, color, head, desc in steps:
        ws.row_dimensions[r].height = 24
        # 説明が長い行は折り返して見切れを防ぐ(C〜J≒全角52字/行)
        dl = max(1, -(-len(desc) // 52))
        ws.row_dimensions[r + 1].height = 14 * dl + 6
        chip(ws, f"B{r}:B{r + 1}", mark, "FFFFFF", color, 16, True, "center")
        style_range(ws, f"B{r}:B{r + 1}", border=Border(left=Side(style="medium", color=color),
                                                        top=Side(style="medium", color=color),
                                                        bottom=Side(style="medium", color=color)))
        ws.merge_cells(f"C{r}:J{r}")
        style_range(ws, f"C{r}:J{r}", font=fnt(11, True, INK), alignment=align("left", "bottom"))
        ws[f"C{r}"] = head
        note(ws, f"C{r + 1}:J{r + 1}", desc, 9, wrap=True)
        r += 3

    r += 1
    ws.row_dimensions[r].height = 22
    chip(ws, f"B{r}:D{r}", "  セルの色のルール", CHIP_NAVY, NAVY)
    r += 1
    ws.row_dimensions[r].height = 20
    chip(ws, f"C{r}:D{r}", "✏️ 黄色 ＝ 入力するセル", F_INPUT, INK, 9.5, False, "center")
    style_range(ws, f"C{r}:D{r}", border=BORDER_INPUT)
    chip(ws, f"F{r}:G{r}", "🔒 グレー ＝ 自動計算", F_AUTO, "5B6472", 9.5, False, "center")
    style_range(ws, f"F{r}:G{r}", border=BORDER_LIGHT)

    r += 2
    ws.row_dimensions[r].height = 22
    chip(ws, f"B{r}:D{r}", "  計算のしくみ", CHIP_NAVY, NAVY)
    r += 1
    note(ws, f"C{r}:J{r}", "購買率 ＝ 参照期間の販売数（CSVの「売上数」列） ÷ 参照期間の動員数合計", 10, INK)
    r += 1
    note(ws, f"C{r}:J{r}", "作る数 ＝ ピーク動員数 × 購買率 × 係数（切り上げ）｜係数＝時間帯係数（商品別の波が出る商品はその実測係数に置き換え）", 10, INK)

    r += 2
    ws.row_dimensions[r].height = 22
    chip(ws, f"B{r}:D{r}", "  注意メモ", CHIP_NAVY, NAVY)
    notes = [
        "・商品の選択・入れ替えは「期間データ」シートの商品名欄（B14〜B33）のプルダウンで行います"
        "（貼られたCSVの商品名から自動でリスト化）。手入力も可能ですが完全一致が必要です。",
        "・ドリンク類・調味料・ＳＥＴ作品コンボ・引換券・包材はプルダウンに出ません"
        "（期間データシート下部の除外リストで自由に変更できます）。",
        "・ピーク動員数には「これから準備する回」の合計動員数を入れます（例：1時間後のピークの回の計）。",
        "・参照期間は3択です：期間A（直近の勢い重視）／期間B（1週間の平均的な売れ方）／期間平均（A+Bを合算した10日間。ブレをならした安全側の基準）。",
        "・時間帯係数は5段階です（朝一0.8倍→昼ピーク1.2倍→夕方1.1倍→夜ピーク1.3倍→レイト0.9倍＋平常。右上の表で編集・"
        "1.2 ＝ 1.2倍 の形で入力）。「④ 商品別の波」を使うと、実測が出る商品はこの全体係数の代わりに商品ごとの"
        "実測係数がかかります（出ない商品は時間帯係数のまま。1商品に両方は掛かりません）。",
        "・貼り付けは必ずオレンジのA4セルを選択して『値の貼り付け』（右クリック→値のみ）。"
        "通常の貼り付けだと色やメモが上書きされます。CSVは各期間1000行まで。",
        "・貼り替える前に、5行目以降のデータだけを選択して削除してください"
        "（1〜4行目の見出し・状態表示は消さないこと）。",
        "・貼り付けシート（CSV貼付A/B・係数貼付①〜④）と係数算出・商品別の波はシート保護済みです（パスワード無し）。"
        "オレンジの貼り付け領域と設定セル以外は書き換えできず、シート全体を選択した貼り付けで内部の計算式が"
        "消える事故もブロックされます（必要な場合は「校閲＞シート保護の解除」で外せます）。",
        "・期間の日付は貼られたCSVの「対象期間」から自動表示されます。貼付後は「期間データ」の"
        "CSV行数・日数チェック（✔／⚠）を確認してください。日数違い・貼付位置ズレ・旧データ残存は⚠が出ます。",
        "・別の期間（連休比較など）を見たいときは、その期間で出力したCSVを貼り替えてください。"
        "日付は自動で切り替わります（期間A=3日・期間B=7日の枠。日付セルは数式のため手入力しないこと）。",
        "・販売数を手入力したい場合は「期間データ」の販売数セルに直接数値を入れても使えます（数式は上書きされます）。",
        "・毎週の作業は「CSV2本の貼り替え」と「動員数の入力」だけです。日付・商品リストは自動で追随します。",
    ]
    for t in notes:
        r += 1
        nl = max(1, -(-int(disp_w(t) * 2) // 100))    # 1行≒全角50字
        ws.row_dimensions[r].height = 15 * nl + 5
        note(ws, f"C{r}:J{r}", t, 9.5, INK, wrap=True)

    r += 2
    ws.row_dimensions[r].height = 22
    chip(ws, f"B{r}:E{r}", "  ⏱ 時間帯係数の較正（月1回・任意）", CHIP_NAVY, NAVY)
    calib_notes = [
        "・「係数貼付①〜④」に、先月の金曜4日分のMSO商品CSV（注文明細・金曜1日分で出力）を1週ずつ貼ると、"
        "「係数算出」シートに時間帯5段階の係数候補が実測から自動で出ます（貼った週だけで平均・4週未満でも可・最大30,000行/シート）。",
        "・算出された候補は「準備数計算」右上のプリセット表のとなり（実測候補列）にも表示されます。"
        "採用するときはプリセット表の係数（J列）へ手で入力してください（自動では書き換わりません）。",
        "・集計ルール：セット親・注文取消・払戻は除外して実個数を数えます。複数日が混ざったCSVは先頭の日付だけを集計し、状態表示に⚠が出ます。",
        "・「商品別の波」シートには、登録商品ごとの時間帯パターン（構成比と商品別係数）が自動で出ます。"
        "準備数計算の「④ 商品別の波」を『使う』にすると、作る数の係数が商品ごとの実測に置き換わります"
        "（データ不足・該当なし・その帯の個数0の商品は自動で全体の時間帯係数。適用値は「商品係数」列で確認できます）。",
        "・時間帯プリセット名の先頭の①〜⑤マークは商品別の波の帯対応キーです。"
        "名前を書き換えるときも先頭のマークは残してください（消すとその時間帯は全体係数になり、警告が出ます）。",
        "・閉店時刻（係数算出シートの時間の区切り）は基本 22:00。レイト営業日は 26:00（＝翌2:00。24時間超え表記OK）"
        "まで対応します。同日閉店・翌日閉店は自動判別。設定と実態のズレ（区切りの外の販売、22時以降の販売が0個の週、"
        "閉店が開店を超える入力など）は係数算出に⚠が出て、商品別の波の適用中は準備数計算の警告にも連動します。",
        "・貼らなくても本体はプリセットの既定係数のまま使えます。",
    ]
    for t in calib_notes:
        r += 1
        nl = max(1, -(-int(disp_w(t) * 2) // 100))    # 1行≒全角50字
        ws.row_dimensions[r].height = 15 * nl + 5
        note(ws, f"C{r}:J{r}", t, 9.5, INK, wrap=True)

    r += 2
    note(ws, f"C{r}:J{r}", "雛形版 v5.2（2026/8）｜数式・レイアウトは自由に調整してください", 8.5)

    # ======================================================== 準備数計算 =====
    ws = wb.create_sheet("準備数計算")
    ws.sheet_properties.tabColor = CORAL
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = 9
    ws.print_area = "A1:H31"

    for c, w in {"A": 2.5, "B": 6, "C": 32, "D": 22, "E": 11, "F": 16,
                 "G": 13, "H": 8, "I": 24, "J": 9, "K": 9, "L": 12, "M": 10}.items():
        ws.column_dimensions[c].width = w
    for c in "LM":                        # ⚙計算用ヘルパー列は隠す(式は動作)
        ws.column_dimensions[c].hidden = True

    ws.row_dimensions[1].height = 34
    title_band(ws, "B1:H1", "　🍿 準備数計算｜ピーク前の仕込み数")
    ws.row_dimensions[2].height = 18
    note(ws, "B2:H2", "参照期間(A/B)の購買率 × ピーク動員数 × 係数（時間帯／商品別の波） で「作る数」を自動計算します", 9)
    ws.row_dimensions[3].height = 6

    # 時間帯プリセット表(編集OK) I3:J10 + 実測候補列K — 5段階の時間帯+基準+空き枠
    chip(ws, "I3:J3", " ⏰ 時間帯プリセット（編集OK）", CHIP_AMBER, INK, 8.5)
    chip(ws, "K3", "実測候補", CHIP_AMBER, GRAY, 8, False, "center")
    for i in range(PRESET_SLOTS):
        rr = 4 + i
        ws.row_dimensions[rr].height = 18
        if i < len(PRESETS):
            ws[f"I{rr}"] = PRESETS[i][0]
            ws[f"J{rr}"] = PRESETS[i][1]
        style_range(ws, f"I{rr}", font=fnt(9, True), fl=fill(F_INPUT),
                    alignment=align("left"), border=BORDER_INPUT)
        # ％を含む表示形式だとExcelの「パーセント自動入力」で 1.2 が 1.2% になるため「倍」表記にする
        style_range(ws, f"J{rr}", font=fnt(9, True, CORAL), fl=fill(F_INPUT),
                    alignment=align("center"), border=BORDER_INPUT, num='0.0"倍"')
        # 実測候補(係数算出シートの転記用①〜⑤を参照表示。採用はJ列へ手入力)。
        # ①〜⑤以外の行は「—」を置き、表の右端が欠けて見えないように揃える
        ws[f"K{rr}"] = f"=係数算出!$L${7 + i}" if i < 5 else "—"
        style_range(ws, f"K{rr}", font=fnt(9, False, GRAY), fl=fill(F_AUTO),
                    alignment=align("center"), border=BORDER_HAIR, num='0.00"倍"')
    ws["K4"].comment = mk_comment("「係数算出」シートで金曜4週分のMSO商品CSVから実測した係数候補です"
                               "（①〜⑤の並び。未貼付のときは「—」）。採用するときは左のJ列へ手で"
                               "入力してください（自動では書き換わりません）。")
    ws["I4"].comment = mk_comment("時間帯の高低差は5段階(朝一→昼ピーク→夕方→夜ピーク→レイト)+基準の平常で"
                               "管理します。名前・時刻の目安・係数とも書き換えでき、空き枠に追加も"
                               "できます(最大7枠)。係数は 1.2 ＝ 1.2倍(×120%) の形で入力してください。"
                               "【重要】名前の先頭の①〜⑤マークは「商品別の波」の帯の対応キーです。"
                               "改名するときも先頭のマークは残してください(消すとその時間帯は全体係数に"
                               "なります)。右の「実測候補」列には係数算出シートの実測値が自動表示されます。")

    # 計算用ヘルパー M4:M9（L/M列は非表示。ラベルは印刷ににじむため置かない）
    helpers = [
        ("L4", "選択期間", "M4", f'=IF(TRIM($D$4)="{SEL_B}",2,IF(TRIM($D$4)="{SEL_A}",1,'
         f'IF(TRIM($D$4)="{SEL_AVG}",3,0)))'),
        ("L5", "動員合計", "M5", "=IF($M$4=2,SUM(期間データ!$C$10:$I$10),IF($M$4=3,"
         "SUM(期間データ!$C$6:$E$6)+SUM(期間データ!$C$10:$I$10),SUM(期間データ!$C$6:$E$6)))"),
        ("L6", "比較動員", "M6", "=IF($M$4=1,SUM(期間データ!$C$10:$I$10),SUM(期間データ!$C$6:$E$6))"),
        ("L7", "時間帯係数", "M7", "=IFERROR(INDEX($J$4:$J$10,MATCH($D$6,$I$4:$I$10,0)),1)"),
        # M8=時間帯の帯番号(選択中プリセット名の先頭マーカー①〜⑤から取得。
        # 表の行位置ベースにするとプリセット表の並べ替えで別の帯の商品別係数が
        # 適用されてしまうため、名前ベースにする。マーカーが無い名前は0=帯対象外)
        ("L8", "帯番号", "M8", '=IF(TRIM($D$6)="",0,'
         'IFERROR(FIND(LEFT(TRIM($D$6),1),"①②③④⑤"),0))'),
        ("L9", "商品別波", "M9", f'=IF(AND(TRIM($D$7)="{WAVE_ON}",$M$8>=1,$M$8<=5,'
         '係数算出!$H$13>0),1,0)'),
    ]
    for _lref, _ltext, vref, formula in helpers:
        ws[vref] = formula
        style_range(ws, vref, font=fnt(8.5, False, GRAY), alignment=align("left"), num="#,##0")
    style_range(ws, "L3:M9", border=BORDER_HAIR)

    # コントロール
    ws.row_dimensions[4].height = 24
    chip(ws, "B4:C4", "  ① 参照期間", CHIP_CORAL, INK, 10)
    ws["D4"] = {"A": SEL_A, "B": SEL_B, "AVG": SEL_AVG}[select]
    style_range(ws, "D4", font=fnt(10.5, True), fl=fill(F_INPUT),
                alignment=align("center"), border=BORDER_INPUT)
    ws.merge_cells("E4:H4")
    ws["E4"] = ('=IF($M$4=3,'
                '"参照: "&IF(期間データ!$C$8="","—",TEXT(期間データ!$C$8,"m/d"))&"〜"&'
                'IF(期間データ!$E$4="","—",TEXT(期間データ!$E$4,"m/d"))&"（期間A+B 平均）",'
                'IF($M$4=2,'
                '"参照: "&IF(期間データ!$C$8="","—",TEXT(期間データ!$C$8,"m/d"))&"〜"&'
                'IF(期間データ!$I$8="","—",TEXT(期間データ!$I$8,"m/d"))&"（期間B 金〜木）",'
                '"参照: "&IF(期間データ!$C$4="","—",TEXT(期間データ!$C$4,"m/d"))&"〜"&'
                'IF(期間データ!$E$4="","—",TEXT(期間データ!$E$4,"m/d"))&"（期間A 直近金土日）"))'
                '&"｜動員合計 "&TEXT($M$5,"#,##0")&"人"')
    style_range(ws, "E4:H4", font=fnt(9.5, False, "5B6472"), alignment=align("left"))

    ws.row_dimensions[5].height = 24
    chip(ws, "B5:C5", "  ② ピーク動員数", CHIP_CORAL, INK, 10)
    if peak is not None:
        ws["D5"] = peak
    ws["D5"].comment = mk_comment("これから準備する回（例：1時間後のピークの回）の合計動員数を"
                               "入力してください。")
    style_range(ws, "D5", font=fnt(10.5, True), fl=fill(F_INPUT),
                alignment=align("center"), border=BORDER_INPUT, num="#,##0")
    note(ws, "E5:H5", "← これから準備する回の合計動員数（例：1時間後のピークの回の計）", 9)

    ws.row_dimensions[6].height = 24
    chip(ws, "B6:C6", "  ③ 時間帯", CHIP_CORAL, INK, 10)
    ws["D6"] = preset
    style_range(ws, "D6", font=fnt(10.5, True), fl=fill(F_INPUT),
                alignment=align("center"), border=BORDER_INPUT)
    ws.merge_cells("E6:H6")
    ws["E6"] = ('="→ 時間帯係数 ×"&TEXT($M$7*100,"0")&"%"&IF($M$9=1,'
                '"（商品別係数が出ない商品に適用）","（右上の表で名前・係数を編集できます）")')
    style_range(ws, "E6:H6", font=fnt(9, False, GRAY), alignment=align("left"))

    # ④ 商品別の波(MSO実測の商品別時間帯係数を作る数に使うか)。
    # 係数は商品ごとに置き換え: 商品係数が出る商品はそれ、出ない商品は時間帯係数。
    # 1商品に両方が同時に掛かることはない
    ws.row_dimensions[7].height = 24
    chip(ws, "B7:C7", "  ④ 商品別の波", CHIP_CORAL, INK, 10)
    ws["D7"] = WAVE_ON
    ws["D7"].comment = mk_comment("『使う』にすると、係数貼付①〜④のMSO実測から算出した商品ごとの"
                               "時間帯係数（商品別の波シート）で作る数を計算します。その商品には"
                               "全体の時間帯係数は掛かりません（置き換え）。データが足りない商品は"
                               "自動で全体の時間帯係数で計算します。『使わない』で全商品一律です。")
    style_range(ws, "D7", font=fnt(10.5, True), fl=fill(F_INPUT),
                alignment=align("center"), border=BORDER_INPUT)
    ws.merge_cells("E7:H7")
    ws["E7"] = ('="→ "&IF($M$9=1,'
                '"商品別係数を適用中 "&SUMPRODUCT(ISNUMBER($H$11:$H$30)*1)&'
                '"/"&SUMPRODUCT(($C$11:$C$30<>"")*1)&"商品（ほかは時間帯係数）",'
                f'IF(TRIM($D$7)="{WAVE_ON}",'
                '"時間帯係数で計算中（MSO未貼付か時間帯が対象外）",'
                '"時間帯係数で計算中（商品別の波オフ）"))')
    style_range(ws, "E7:H7", font=fnt(9, False, GRAY), alignment=align("left"))
    ws.row_dimensions[8].height = 6

    ws.row_dimensions[9].height = 26
    ws.merge_cells("B9:H9")
    ws["B9"] = ('=TRIM('
                'IF($M$4=0,"⚠ 参照期間の選択が不正です（現在は期間A扱い）。リストから選び直してください。","")&" "&'
                'IF(OR(AND($M$4<>2,CSV貼付A!$N$5=""),AND($M$4<>1,CSV貼付B!$N$5="")),'
                '"⚠ 参照に必要なCSVが未貼付です（CSV貼付"&IF(AND($M$4<>2,CSV貼付A!$N$5=""),"A","")&IF(AND($M$4<>1,CSV貼付B!$N$5=""),"B","")&"）。","")&" "&'
                'IF(OR(AND($M$4<>2,ISNUMBER(SEARCH("⚠",期間データ!$G$4))),'
                'AND($M$4<>1,ISNUMBER(SEARCH("⚠",期間データ!$B$11)))),'
                '"⚠ 選択した期間のCSVに問題があります（期間データシートの表示を確認）。","")&" "&'
                'IF(OR(AND($M$4<>2,COUNT(期間データ!$C$6:$E$6)<3),'
                'AND($M$4<>1,COUNT(期間データ!$C$10:$I$10)<7)),'
                '"⚠ 選択した期間の動員数がそろっていません。","")&" "&'
                'IF(AND(IF($M$4=2,COUNT(期間データ!$C$10:$I$10)=7,'
                'IF($M$4=3,AND(COUNT(期間データ!$C$6:$E$6)=3,COUNT(期間データ!$C$10:$I$10)=7),'
                'COUNT(期間データ!$C$6:$E$6)=3)),$M$5<=0),'
                '"⚠ 参照期間の動員数合計が0です（期間データシートを確認）。","")&" "&'
                'IF(SUMPRODUCT((期間データ!$B$14:$B$33<>"")*'
                '(COUNTIF(期間データ!$B$14:$B$33,期間データ!$B$14:$B$33)>1))>0,'
                '"⚠ 商品名が重複しています（集計が二重になります）。","")&" "&'
                'IF(AND(IF($M$4=3,OR(CSV貼付A!$N$5<>"",CSV貼付B!$N$5<>""),'
                'IF($M$4=2,CSV貼付B!$N$5<>"",CSV貼付A!$N$5<>"")),'
                'SUMPRODUCT((期間データ!$B$14:$B$33<>"")*'
                f'(IF($M$4=2,COUNTIF(CSV貼付B!$N$5:$N${CSV_END},期間データ!$B$14:$B$33),'
                f'IF($M$4=3,COUNTIF(CSV貼付A!$N$5:$N${CSV_END},期間データ!$B$14:$B$33)+'
                f'COUNTIF(CSV貼付B!$N$5:$N${CSV_END},期間データ!$B$14:$B$33),'
                f'COUNTIF(CSV貼付A!$N$5:$N${CSV_END},期間データ!$B$14:$B$33)))=0))>0),'
                '"⚠ 参照期間のCSVに無い商品名があります（別期間のみの商品か、表記を確認。期間販売数0扱い）。","")&" "&'
                'IF(COUNTIF($D$11:$D$30,"<0")>0,'
                '"⚠ 期間販売数がマイナスの商品があります（返品超過）。作る数は0扱いです。","")&" "&'
                'IF(AND($D$6<>"",ISNA(MATCH($D$6,$I$4:$I$10,0))),'
                '"⚠ 時間帯プリセット名が表にありません（係数100%扱い）。","")&" "&'
                'IF(OR($D$5="",$D$5=0,NOT(ISNUMBER($D$5))),"⚠ ピーク動員数が未入力か数値ではありません。","")&" "&'
                'IF($M$7=0,"⚠ 時間帯係数が0です。","")&" "&'
                'IF(AND($M$7>0,OR($M$7<0.5,$M$7>5)),'
                '"⚠ 時間帯係数が×50%〜×500%の範囲外です。プリセット表の係数を確認してください。","")&" "&'
                f'IF(AND($D$7<>"",TRIM($D$7)<>"{WAVE_ON}",TRIM($D$7)<>"{WAVE_OFF}"),'
                '"⚠ 商品別の波の設定が不正です（使わない扱い）。","")&" "&'
                f'IF(AND(TRIM($D$7)="{WAVE_ON}",係数算出!$H$13>0,$M$8=0,'
                'SUMPRODUCT((TRIM($I$4:$I$10)<>"")*'
                'ISNUMBER(FIND(LEFT(TRIM($I$4:$I$10),1),"①②③④⑤")))<5),'
                '"⚠ 時間帯名の先頭に①〜⑤のマークが無いため商品別の波を適用できません'
                '（全体係数で計算中。プリセット名の先頭マークは消さないでください）。","")&" "&'
                'IF(AND($M$9=1,OR('
                'AND(係数算出!$H$4>係数算出!$C$4,係数算出!$H$4<係数算出!$G$4),'
                'AND(係数算出!$H$4>=1,MOD(係数算出!$H$4,1)>係数算出!$C$4),'
                '係数算出!$N$5>0,係数算出!$O$5>0)),'
                '"⚠ 係数算出シートに時間の区切りの警告があります'
                '（商品別係数が不正確な可能性。係数算出シートを確認してください）。",""))')
    style_range(ws, "B9:H9", font=fnt(8.5, True, RED), alignment=align("left", wrap=True))

    # 表ヘッダー
    ws.row_dimensions[10].height = 34
    for ref, text in [("B10", "No."), ("C10", "商品名"), ("D10", "期間販売数"),
                      ("E10", "購買率"), ("G10", "（参考）\n比較期間"),
                      ("H10", "商品係数\n(自動)")]:
        style_range(ws, ref, font=fnt(9.5, True, "FFFFFF"), fl=fill(NAVY),
                    alignment=align("center", "center", True), border=BORDER_LIGHT)
        ws[ref] = text
    ws["C10"].comment = mk_comment("商品の選択・入れ替えは「期間データ」シートの商品名欄"
                                "(B14〜B33)のプルダウンで行ってください。ここは自動表示です。")
    style_range(ws, "F10", font=fnt(11, True, "FFFFFF"), fl=fill(CORAL),
                alignment=align("center", "center", True), border=BORDER_LIGHT)
    ws["F10"] = "👉 作る数\n(この数を準備)"

    for i in range(N_SLOTS):
        r = ROW_M0 + i
        dr = ROW_P0 + i
        ws.row_dimensions[r].height = 21
        ws[f"B{r}"] = i + 1
        ws[f"C{r}"] = f'=IF(期間データ!B{dr}="","",期間データ!B{dr})'
        ws[f"D{r}"] = (f'=IF($C{r}="","",IF($M$4=2,期間データ!D{dr},'
                       f'IF($M$4=3,期間データ!C{dr}+期間データ!D{dr},期間データ!C{dr})))')
        ws[f"E{r}"] = f'=IF($C{r}="","",IF($M$5<=0,"要確認",D{r}/$M$5))'
        # 係数は商品ごとに置き換え: 商品別の波が使える商品(M9=1かつ商品係数が正の数値)は
        # その係数、出ない商品は全体の時間帯係数M7。1商品に両方が同時に掛かることはない
        wr = WAVE_ROW0 + i
        coef_p = f"INDEX({WAVE_SHEET}!$I${wr}:$M${wr},1,$M$8)"
        # 係数<=0はM7へ(波シート側でも"—"にしているが、0が数値として掛かり
        # 作る数が無警告で0になる事故への二重ガード)
        eff = f'IF(AND($M$9=1,ISNUMBER({coef_p}),IFERROR({coef_p},0)>0),{coef_p},$M$7)'
        ws[f"F{r}"] = (f'=IF(OR($C{r}="",NOT(ISNUMBER($D$5))),"",'
                       f'IF(ISNUMBER($E{r}),'
                       f'MAX(0,ROUNDUP($D$5*$E{r}*{eff},0)),"—"))')
        ws[f"G{r}"] = (f'=IF($C{r}="","",'
                       f'IF(ISNUMBER(SEARCH("⚠",IF($M$4=1,期間データ!$B$11,期間データ!$G$4))),"要確認",'
                       f'IF($M$6<=0,"－",IF($M$4=1,期間データ!D{dr},期間データ!C{dr})/$M$6)))')
        # {coef_p}はM8=0のときINDEX列0でエラーになる。ORは引数のエラーを伝播する
        # ため、比較はIFERRORで包む(ISNUMBERはエラーを吸収するのでそのままでよい)
        ws[f"H{r}"] = (f'=IF($C{r}="","",'
                       f'IF(OR($M$9=0,NOT(ISNUMBER({coef_p})),IFERROR({coef_p},0)<=0),"—",{coef_p}))')
        style_range(ws, f"B{r}", font=fnt(9, False, GRAY), alignment=align("center"))
        style_range(ws, f"C{r}", font=fnt(10.5), alignment=align("left"))
        style_range(ws, f"D{r}", font=fnt(10, False, "5B6472"), alignment=align("center"), num="#,##0")
        style_range(ws, f"E{r}", font=fnt(10, False, "5B6472"), alignment=align("center"), num="0.0%")
        style_range(ws, f"F{r}", font=fnt(13, True, CORAL), fl=fill(F_BASE),
                    alignment=align("center"), num="#,##0")
        style_range(ws, f"G{r}", font=fnt(9, False, GRAY), alignment=align("center"), num="0.0%")
        style_range(ws, f"H{r}", font=fnt(9, False, GRAY), alignment=align("center"), num='0.00"倍"')
        if i % 2:
            for col in "BCDEGH":
                ws[f"{col}{r}"].fill = fill(F_ZEBRA)
        for col in "BCDEGH":
            ws[f"{col}{r}"].border = Border(bottom=hair, left=hair, right=hair)
        ws[f"F{r}"].border = Border(bottom=hair, left=coral_side, right=coral_side)

    last = ROW_M0 + N_SLOTS - 1
    ws[f"F{last}"].border = Border(bottom=coral_side, left=coral_side, right=coral_side)
    ws.row_dimensions[last + 1].height = 18
    ws.row_dimensions[last + 1].height = 30
    note(ws, f"B{last + 1}:H{last + 1}",
         "※ 作る数 ＝ ピーク動員数 × 購買率 × 係数（切り上げ）｜係数 ＝ 商品係数（右列。商品別の波の実測）が"
         "あればそれ、「—」の商品は時間帯係数（1商品に両方は掛かりません）｜"
         "参考列 ＝ A選択時は期間B、それ以外は期間A", 8.5, wrap=True)

    bar = DataBarRule(start_type="num", start_value=0, end_type="max", color=CORAL, showValue=True)
    ws.conditional_formatting.add(f"F{ROW_M0}:F{last}", bar)
    rate_warn = FormulaRule(formula=[f"ISTEXT(E{ROW_M0})"],
                            font=Font(name=FONT_NAME, size=9, bold=True, color=RED))
    ws.conditional_formatting.add(f"E{ROW_M0}:E{last}", rate_warn)
    ref_warn = FormulaRule(formula=[f'ISNUMBER(SEARCH("要確認",G{ROW_M0}))'],
                           font=Font(name=FONT_NAME, size=9, bold=True, color=RED))
    ws.conditional_formatting.add(f"G{ROW_M0}:G{last}", ref_warn)

    dv_period = DataValidation(type="list", formula1=f'"{SEL_A},{SEL_B},{SEL_AVG}"',
                               allow_blank=False, showErrorMessage=True)
    dv_period.error = f"リストから選んでください（{SEL_A}／{SEL_B}／{SEL_AVG}）"
    dv_period.errorTitle = "参照期間"
    ws.add_data_validation(dv_period)
    dv_period.add("D4")

    dv_peak = DataValidation(type="whole", operator="between", formula1="0", formula2="999999",
                             showErrorMessage=True)
    dv_peak.error = "ピーク動員数は 0〜999,999 の整数で入力してください"
    dv_peak.errorTitle = "ピーク動員数"
    ws.add_data_validation(dv_peak)
    dv_peak.add("D5")

    dv_preset = DataValidation(type="list", formula1="=$I$4:$I$10", allow_blank=True,
                               showErrorMessage=False)
    ws.add_data_validation(dv_preset)
    dv_preset.add("D6")

    dv_wave = DataValidation(type="list", formula1=f'"{WAVE_ON},{WAVE_OFF}"',
                             allow_blank=True, showErrorMessage=True)
    dv_wave.error = f"「{WAVE_ON}」か「{WAVE_OFF}」を選んでください"
    dv_wave.errorTitle = "商品別の波"
    ws.add_data_validation(dv_wave)
    dv_wave.add("D7")

    ws.freeze_panes = "A11"

    # ========================================================== 印刷用 =======
    ws = wb.create_sheet("印刷用")
    ws.sheet_properties.tabColor = GREEN
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = 9
    ws.print_area = "A1:F28"

    for c, w in {"A": 2.5, "B": 7, "C": 32, "D": 16, "E": 13, "F": 2.5}.items():
        ws.column_dimensions[c].width = w

    ws.row_dimensions[1].height = 36
    title_band(ws, "B1:E1", "　🍿 仕込み指示書（コンセッション）")
    ws.row_dimensions[2].height = 24
    chip(ws, "B2:E2", "  👇 この数を作ってください", CHIP_CORAL, CORAL, 12, True)

    ws.row_dimensions[3].height = 22
    ws.merge_cells("B3:E3")
    ws["B3"] = ('="参照期間: "&IF(準備数計算!$M$4=3,"期間平均（A+B）",'
                'IF(準備数計算!$M$4=2,"期間B（金〜木）","期間A（直近金土日）"))&" "&'
                'IF(準備数計算!$M$4=1,'
                'IF(期間データ!$C$4="","",TEXT(期間データ!$C$4,"m/d")&"〜"&TEXT(期間データ!$E$4,"m/d")),'
                'IF(期間データ!$C$8="","",TEXT(期間データ!$C$8,"m/d")&"〜"&'
                'IF(準備数計算!$M$4=2,TEXT(期間データ!$I$8,"m/d"),TEXT(期間データ!$E$4,"m/d"))))&'
                '" ｜ ピーク動員数: "&IF(準備数計算!$D$5="","（未入力）",'
                'TEXT(準備数計算!$D$5,"#,##0")&"人")')
    style_range(ws, "B3:E3", font=fnt(11, True, INK), alignment=align("left"))
    ws.row_dimensions[4].height = 20
    ws.merge_cells("B4:E4")
    ws["B4"] = ('="時間帯: "&準備数計算!$D$6&"（時間帯係数 ×"&TEXT(準備数計算!$M$7*100,"0")&"%）"&'
                'IF(準備数計算!$M$9=1,"　｜　商品別の波 適用中","")')
    style_range(ws, "B4:E4", font=fnt(9.5, False, "5B6472"), alignment=align("left"))
    ws.row_dimensions[5].height = 22
    note(ws, "B5:E5", "日付・回：＿＿＿＿＿＿＿＿＿＿　　作成者：＿＿＿＿＿＿　　確認者：＿＿＿＿＿＿", 10, INK)
    ws.row_dimensions[6].height = 14
    ws.merge_cells("B6:E6")
    ws["B6"] = "=準備数計算!$B$9"
    style_range(ws, "B6:E6", font=fnt(8, True, RED), alignment=align("left"))

    ws.row_dimensions[7].height = 26
    for ref, text in [("B7", "No."), ("C7", "商品名"), ("E7", "できたら✓")]:
        style_range(ws, ref, font=fnt(10, True, "FFFFFF"), fl=fill(NAVY),
                    alignment=align("center"), border=BORDER_LIGHT)
        ws[ref] = text
    style_range(ws, "D7", font=fnt(11, True, "FFFFFF"), fl=fill(CORAL),
                alignment=align("center"), border=BORDER_LIGHT)
    ws["D7"] = "作る数"

    for i in range(N_SLOTS):
        r = 8 + i
        mr = ROW_M0 + i
        ws.row_dimensions[r].height = 24
        ws[f"B{r}"] = f'=IF(準備数計算!$C{mr}="","",{i + 1})'
        ws[f"C{r}"] = f'=IF(準備数計算!$C{mr}="","",準備数計算!$C{mr})'
        ws[f"D{r}"] = f'=IF(準備数計算!$C{mr}="","",準備数計算!$F{mr})'
        ws[f"E{r}"] = f'=IF(準備数計算!$C{mr}="","","☐")'
        style_range(ws, f"B{r}", font=fnt(9.5, False, GRAY), alignment=align("center"))
        style_range(ws, f"C{r}", font=fnt(11), alignment=align("left"))
        style_range(ws, f"D{r}", font=fnt(14, True, CORAL), fl=fill(F_BASE),
                    alignment=align("center"), num="#,##0")
        style_range(ws, f"E{r}", font=fnt(12, False, "B9C0CC"), alignment=align("center"))
        for col in "BCE":
            ws[f"{col}{r}"].border = BORDER_LIGHT
        ws[f"D{r}"].border = Border(bottom=thin, top=thin, left=coral_side, right=coral_side)

    ws.row_dimensions[28].height = 16
    note(ws, "B28:E28", "※ 数字は「準備数計算」シートから自動で入ります｜A4縦・1ページ印刷", 8)

    # ======================================================== 期間データ =====
    ws = wb.create_sheet("期間データ")
    ws.sheet_properties.tabColor = TEAL
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = 9
    ws.print_area = "A1:K34"

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 30
    for c in "CDEFGHI":
        ws.column_dimensions[c].width = 10.5
    ws.column_dimensions["J"].width = 11
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 2.5
    for c, w in {"M": 6, "N": 28, "O": 6, "P": 28, "Q": 28}.items():
        ws.column_dimensions[c].width = w
        ws.column_dimensions[c].hidden = True

    ws.row_dimensions[1].height = 34
    title_band(ws, "A1:K1", "　📅 期間データ（動員数・販売実績）")
    ws.row_dimensions[2].height = 20
    # 凡例チップは1列空けて分離(密着させると1つの帯が色変わりして見える)
    note(ws, "A2", "凡例:", 8.5, GRAY, h="right")
    chip(ws, "B2", "✏️ 入力セル", F_INPUT, INK, 8.5, False, "center")
    style_range(ws, "B2", border=BORDER_INPUT)
    chip(ws, "D2:E2", "🔒 自動計算", F_AUTO, "5B6472", 8.5, False, "center")
    style_range(ws, "D2:E2", border=BORDER_LIGHT)
    note(ws, "F2:K2", "販売数はCSV貼付A/Bから自動集計。商品名はプルダウンで選択。", 8.5)
    ws.row_dimensions[3].height = 6

    # 期間A(直近金土日) — 日付はCSVの対象期間から自動表示
    pa_s = parse_ymd("CSV貼付A!$D$5")
    pa_e = parse_ymd("CSV貼付A!$E$5")
    ws.row_dimensions[4].height = 22
    chip(ws, "B4", "  期間A（直近金土日）｜日付は自動", CHIP_TEAL, INK, 10)
    ws.row_dimensions[5].height = 16
    note(ws, "B5", "  曜日", 8.5, GRAY)
    ws.row_dimensions[6].height = 22
    chip(ws, "B6", "  動員数（人）", CHIP_TEAL, INK, 10)
    ws["C4"] = f'=IF(CSV貼付A!$N$5="","",IF({pa_s}=-1,"",{pa_s}))'
    ws["D4"] = '=IF(OR($C$4="",$E$4="",$E$4<$C$4),"",$C$4+1)'
    ws["E4"] = f'=IF(CSV貼付A!$N$5="","",IF({pa_e}=-1,"",{pa_e}))'
    for j in range(3):
        col = get_column_letter(3 + j)
        style_range(ws, f"{col}4", font=fnt(10, True), fl=fill(F_AUTO),
                    alignment=align("center"), border=BORDER_LIGHT, num="m/d")
        ws[f"{col}5"] = f'=IF({col}$4="","",CHOOSE(WEEKDAY({col}$4),{WEEKDAY_JA}))'
        style_range(ws, f"{col}5", font=fnt(8.5, False, GRAY), alignment=align("center"))
        if att_a:
            ws[f"{col}6"] = att_a[j]
        style_range(ws, f"{col}6", font=fnt(10.5), fl=fill(F_INPUT),
                    alignment=align("center"), border=BORDER_INPUT, num="#,##0")
    ws["J6"] = '=IF(COUNT(C6:E6)=0,"",SUM(C6:E6))'
    style_range(ws, "J6", font=fnt(10.5, True, TEAL), fl=fill(F_AUTO),
                alignment=align("center"), border=BORDER_LIGHT, num="#,##0")
    note(ws, "J5", "A計", 8, GRAY, h="center")
    ws.merge_cells("G4:K4")
    ws["G4"] = ('=IF(CSV貼付A!$N$5="","（CSV貼付Aにデータを貼り付けてください）",'
                '"CSV貼付A: "&COUNTA(CSV貼付A!$N$5:$N$' + str(CSV_END) + ')&"行｜"&'
                'IF(CSV貼付A!$N$4<>"商品名","⚠ 貼付位置がずれています（ヘッダーごとならA4、データのみならA5から）",'
                'IF(CSV貼付A!$A$5="タイトル","⚠ 貼付開始セルがずれています（A4から貼り直してください）",'
                'IF(OR($C$4="",$E$4=""),"⚠ 対象期間を読み取れません（CSVのD/E列を確認）",'
                'IF($E$4<$C$4,"⚠ 対象期間が逆転しています（CSVのD/E列を確認）",'
                'TRIM('
                'IF($E$4-$C$4<>2,"⚠ "&($E$4-$C$4+1)&"日分のCSVです（期間Aは金土日3日想定）",'
                '"✔ "&($E$4-$C$4+1)&"日間")&" "&'
                'IF(SUMPRODUCT((CSV貼付A!$D$5:$D$' + str(CSV_END) + '<>"")*(CSV貼付A!$D$5:$D$' + str(CSV_END) + '<>CSV貼付A!$D$5))+'
                'SUMPRODUCT((CSV貼付A!$E$5:$E$' + str(CSV_END) + '<>"")*(CSV貼付A!$E$5:$E$' + str(CSV_END) + '<>CSV貼付A!$E$5))>0,'
                '"⚠ 別期間の行が混ざっています（前回分を削除して貼り直し）","")&" "&'
                'IF(WEEKDAY($C$4)<>6,"※開始が金曜以外","")&" "&'
                'IF(TODAY()-$E$4>9,"⚠ 古いデータの可能性（終了日が"&(TODAY()-$E$4)&"日前）","")'
                '))))))')
    style_range(ws, "G4:K4", font=fnt(9, False, "5B6472"), alignment=align("left"))
    ws["C4"].comment = mk_comment("貼られたCSVの対象期間から自動表示されます（入力不要）。")

    ws.row_dimensions[7].height = 6
    # 期間B(前週 金〜木) — 日付はCSVの対象期間から自動表示
    pb_s = parse_ymd("CSV貼付B!$D$5")
    pb_e = parse_ymd("CSV貼付B!$E$5")
    ws.row_dimensions[8].height = 22
    chip(ws, "B8", "  期間B（前週 金〜木）｜日付は自動", CHIP_TEAL, INK, 10)
    ws.row_dimensions[9].height = 16
    note(ws, "B9", "  曜日", 8.5, GRAY)
    ws.row_dimensions[10].height = 22
    chip(ws, "B10", "  動員数（人）", CHIP_TEAL, INK, 10)
    ws["C8"] = f'=IF(CSV貼付B!$N$5="","",IF({pb_s}=-1,"",{pb_s}))'
    for j in range(1, 6):
        ws[f"{get_column_letter(3 + j)}8"] = f'=IF(OR($C$8="",$I$8="",$I$8<$C$8),"",$C$8+{j})'
    ws["I8"] = f'=IF(CSV貼付B!$N$5="","",IF({pb_e}=-1,"",{pb_e}))'
    for j in range(7):
        col = get_column_letter(3 + j)
        style_range(ws, f"{col}8", font=fnt(10, True), fl=fill(F_AUTO),
                    alignment=align("center"), border=BORDER_LIGHT, num="m/d")
        ws[f"{col}9"] = f'=IF({col}$8="","",CHOOSE(WEEKDAY({col}$8),{WEEKDAY_JA}))'
        style_range(ws, f"{col}9", font=fnt(8.5, False, GRAY), alignment=align("center"))
        if att_b:
            ws[f"{col}10"] = att_b[j]
        style_range(ws, f"{col}10", font=fnt(10.5), fl=fill(F_INPUT),
                    alignment=align("center"), border=BORDER_INPUT, num="#,##0")
    ws["J10"] = '=IF(COUNT(C10:I10)=0,"",SUM(C10:I10))'
    style_range(ws, "J10", font=fnt(10.5, True, TEAL), fl=fill(F_AUTO),
                alignment=align("center"), border=BORDER_LIGHT, num="#,##0")
    note(ws, "J9", "B計", 8, GRAY, h="center")
    ws.row_dimensions[11].height = 16
    ws.merge_cells("B11:K11")
    ws["B11"] = ('=IF(CSV貼付B!$N$5="","CSV貼付B:（未貼付）",'
                 '"CSV貼付B: "&COUNTA(CSV貼付B!$N$5:$N$' + str(CSV_END) + ')&"行｜"&'
                 'IF(CSV貼付B!$N$4<>"商品名","⚠ 貼付位置がずれています（ヘッダーごとならA4、データのみならA5から）",'
                 'IF(CSV貼付B!$A$5="タイトル","⚠ 貼付開始セルがずれています（A4から貼り直してください）",'
                 'IF(OR($C$8="",$I$8=""),"⚠ 対象期間を読み取れません（CSVのD/E列を確認）",'
                 'IF($I$8<$C$8,"⚠ 対象期間が逆転しています（CSVのD/E列を確認）",'
                 'TRIM('
                 'IF($I$8-$C$8<>6,"⚠ "&($I$8-$C$8+1)&"日分のCSVです（期間Bは金〜木7日想定）",'
                 '"✔ "&($I$8-$C$8+1)&"日間")&" "&'
                 'IF(SUMPRODUCT((CSV貼付B!$D$5:$D$' + str(CSV_END) + '<>"")*(CSV貼付B!$D$5:$D$' + str(CSV_END) + '<>CSV貼付B!$D$5))+'
                 'SUMPRODUCT((CSV貼付B!$E$5:$E$' + str(CSV_END) + '<>"")*(CSV貼付B!$E$5:$E$' + str(CSV_END) + '<>CSV貼付B!$E$5))>0,'
                 '"⚠ 別期間の行が混ざっています（前回分を削除して貼り直し）","")&" "&'
                 'IF(WEEKDAY($C$8)<>6,"※開始が金曜以外","")&" "&'
                 'IF(TODAY()-$I$8>9,"⚠ 古いデータの可能性（終了日が"&(TODAY()-$I$8)&"日前）","")'
                 '))))))')
    style_range(ws, "B11:K11", font=fnt(9, False, "5B6472"), alignment=align("left"))

    # 週末色分け
    for rng, drow in [("C4:E5", 4), ("C8:I9", 8)]:
        sun = FormulaRule(formula=[f'WEEKDAY(C${drow})=1'],
                          font=Font(name=FONT_NAME, color=RED, bold=True))
        sat = FormulaRule(formula=[f'WEEKDAY(C${drow})=7'],
                          font=Font(name=FONT_NAME, color="3B6FD4", bold=True))
        ws.conditional_formatting.add(rng, sun)
        ws.conditional_formatting.add(rng, sat)

    ws.row_dimensions[12].height = 6
    ws.row_dimensions[13].height = 30
    ws.merge_cells("E13:K13")
    for ref, text in [("A13", "No."), ("B13", "商品名（プルダウンで選択）"),
                      ("C13", "期間A\n販売数"), ("D13", "期間B\n販売数"),
                      ("E13:K13", "メモ（自由記入）")]:
        style_range(ws, ref, font=fnt(9.5, True, "FFFFFF"), fl=fill(NAVY),
                    alignment=align("center", "center", True), border=BORDER_LIGHT)
        ws[ref.split(":")[0]] = text

    for i in range(N_SLOTS):
        r = ROW_P0 + i
        ws.row_dimensions[r].height = 20
        ws[f"A{r}"] = i + 1
        style_range(ws, f"A{r}", font=fnt(9, False, GRAY), alignment=align("center"))
        if i < len(products):
            ws[f"B{r}"] = products[i]
        style_range(ws, f"B{r}", font=fnt(10), fl=fill(F_INPUT),
                    alignment=align("left"), border=BORDER_INPUT)
        ws[f"C{r}"] = (f'=IF($B{r}="","",SUMIF(CSV貼付A!$N$5:$N${CSV_END},$B{r},'
                       f'CSV貼付A!$AA$5:$AA${CSV_END}))')
        ws[f"D{r}"] = (f'=IF($B{r}="","",SUMIF(CSV貼付B!$N$5:$N${CSV_END},$B{r},'
                       f'CSV貼付B!$AA$5:$AA${CSV_END}))')
        for col, zebra in (("C", i % 2), ("D", i % 2)):
            style_range(ws, f"{col}{r}", font=fnt(10, False, "5B6472"),
                        fl=fill(F_ZEBRA if zebra else F_AUTO),
                        alignment=align("center"), num="#,##0")
        ws.merge_cells(f"E{r}:K{r}")
        style_range(ws, f"E{r}:K{r}", font=fnt(9), alignment=align("left"))
        for col in "ABCDEFGHIJK":
            ws[f"{col}{r}"].border = Border(bottom=hair, left=hair, right=hair)
        ws[f"B{r}"].border = BORDER_INPUT

    ws[f"B{ROW_P0}"].comment = mk_comment("商品は最大20枠(B14〜B33)まで登録できます。プルダウンには"
                                       "CSV貼付A/Bの商品名が自動で並びます(項目数の上限はありません)。"
                                       "ドリンク・包材などはシート下部の除外リストで検索対象外です。"
                                       "手入力する場合はCSVの商品名と完全一致させてください。")
    dup_rule = FormulaRule(
        formula=[f'AND($B{ROW_P0}<>"",COUNTIF($B${ROW_P0}:$B${ROW_P0 + N_SLOTS - 1},$B{ROW_P0})>1)'],
        font=Font(name=FONT_NAME, bold=True, color=RED), fill=fill("FDECEC"))
    ws.conditional_formatting.add(f"B{ROW_P0}:B{ROW_P0 + N_SLOTS - 1}", dup_rule)

    dv_att = DataValidation(type="whole", operator="between", formula1="0", formula2="999999",
                            showErrorMessage=True)
    dv_att.error = "動員数は 0〜999,999 の整数で入力してください（文字や記号は不可）"
    dv_att.errorTitle = "動員数"
    ws.add_data_validation(dv_att)
    dv_att.add("C6:E6")
    dv_att.add("C10:I10")

    dv_name = DataValidation(type="list", formula1="商品リスト", allow_blank=True,
                             showErrorMessage=False)
    ws.add_data_validation(dv_name)
    dv_name.add(f"B{ROW_P0}:B{ROW_P0 + N_SLOTS - 1}")

    last = ROW_P0 + N_SLOTS - 1
    ws.row_dimensions[last + 1].height = 26
    note(ws, f"A{last + 1}:K{last + 1}",
         "※ 販売数のセルは自動計算です（CSVの「売上数」列を商品名で集計）。CSVを使わず手入力したい場合は、"
         "数値を直接入力しても使えます（自動集計に戻すには、同じ列の上下のセルの数式をコピーしてください。"
         "となりの列はもう一方の期間を参照しているため使わないでください）。",
         8.5, GRAY, wrap=True)

    # プルダウン検索から除外する小分類(編集OK)
    ws.row_dimensions[EXC_TOP - 1].height = 22
    chip(ws, f"B{EXC_TOP - 1}:E{EXC_TOP - 1}", "  🔎 プルダウン検索から除外する小分類（編集OK）",
         CHIP_NAVY, NAVY, 9.5)
    for i in range(EXC_SLOTS):
        r = EXC_TOP + i
        ws.row_dimensions[r].height = 18
        if i < len(EXCLUDE_CATS):
            ws[f"B{r}"] = EXCLUDE_CATS[i]
        style_range(ws, f"B{r}", font=fnt(9.5), fl=fill(F_INPUT),
                    alignment=align("left"), border=BORDER_INPUT)
        # CSVに実在する小分類名かの照合(タイポ・表記ゆれの空振り検知)
        ws[f"C{r}"] = (f'=IF(OR(TRIM($B{r})="",AND(CSV貼付A!$N$5="",CSV貼付B!$N$5="")),"",'
                       f'IF(COUNTIF(CSV貼付A!$H$5:$H${CSV_END},$B{r})+'
                       f'COUNTIF(CSV貼付B!$H$5:$H${CSV_END},$B{r})=0,"⚠該当なし","✔"))')
        style_range(ws, f"C{r}", font=fnt(8.5, False, GRAY), alignment=align("center"))
    exc_ok = FormulaRule(formula=[f'C{EXC_TOP}="✔"'],
                         font=Font(name=FONT_NAME, size=8.5, color=GREEN))
    exc_ng = FormulaRule(formula=[f'ISNUMBER(SEARCH("⚠",C{EXC_TOP}))'],
                         font=Font(name=FONT_NAME, size=8.5, bold=True, color=RED))
    ws.conditional_formatting.add(f"C{EXC_TOP}:C{EXC_END}", exc_ok)
    ws.conditional_formatting.add(f"C{EXC_TOP}:C{EXC_END}", exc_ng)
    note(ws, f"D{EXC_TOP}:K{EXC_TOP + 3}",
         "← ここに書いた小分類（CSVのH列「小分類名」と完全一致）の商品は、商品名プルダウンに"
         "出なくなります。集計からは除外されないため、手入力すれば集計できます。"
         "不要になった名前はセルの値をDeleteで消し、追加は空き枠に入力してください（最大12件）。"
         "行そのものの挿入・削除はしないでください（内部の自動計算が壊れます）。"
         "左の✔/⚠は、CSVに実在する小分類名かの照合結果です。", 8.5, GRAY, wrap=True)
    ws[f"B{EXC_TOP}"].comment = mk_comment("既定: ドリンク類(コールド/コーヒー/アルコール/その他ドリンク/ホット)、"
                                        "調味料類、ＳＥＴ作品コンボ、引換券、コンセ包材。")

    # 商品リスト抽出ヘルパー(非表示列 M〜Q)
    for ref, text in [("M4", "⚙A順"), ("N4", "⚙Aリスト"), ("O4", "⚙B順"),
                      ("P4", "⚙Bリスト"), ("Q4", "⚙商品リスト")]:
        note(ws, ref, text, 8, GRAY)
    exc = f"$B${EXC_TOP}:$B${EXC_END}"
    for i in range(CSV_MAX):
        r = 5 + i
        # 重複判定は非除外カテゴリの先行行のみ数える(除外行で先に出た同名商品が消えないように)
        ws[f"M{r}"] = (f'=IF(CSV貼付A!$N{r}="","",'
                       f'IF(COUNTIF({exc},CSV貼付A!$H{r})>0,"",'
                       f'IF(SUMPRODUCT((CSV貼付A!$N$5:$N{r}=CSV貼付A!$N{r})*'
                       f'(COUNTIF({exc},CSV貼付A!$H$5:$H{r})=0))>1,"",'
                       f'MAX($M$4:M{r - 1})+1)))')
        ws[f"O{r}"] = (f'=IF(CSV貼付B!$N{r}="","",'
                       f'IF(COUNTIF({exc},CSV貼付B!$H{r})>0,"",'
                       f'IF(SUMPRODUCT((CSV貼付B!$N$5:$N{r}=CSV貼付B!$N{r})*'
                       f'(COUNTIF({exc},CSV貼付B!$H$5:$H{r})=0))>1,"",'
                       f'MAX($O$4:O{r - 1})+1)))')
    for i in range(LIST_MAX):
        r = 5 + i
        ws[f"N{r}"] = (f'=IFERROR(INDEX(CSV貼付A!$N$5:$N${CSV_END},MATCH(ROW()-4,$M$5:$M${CSV_END},0)),"")')
        ws[f"P{r}"] = (f'=IFERROR(INDEX(CSV貼付B!$N$5:$N${CSV_END},MATCH(ROW()-4,$O$5:$O${CSV_END},0)),"")')
    for i in range(2 * LIST_MAX):
        r = 5 + i
        ws[f"Q{r}"] = (f'=IF(ROW()-4<=MAX($M$5:$M${CSV_END}),'
                       f'IFERROR(INDEX($N$5:$N${4 + LIST_MAX},ROW()-4),""),'
                       f'IFERROR(INDEX($P$5:$P${4 + LIST_MAX},ROW()-4-MAX($M$5:$M${CSV_END})),""))')

    wb.defined_names["商品リスト"] = DefinedName(
        "商品リスト",
        attr_text=(f"OFFSET(期間データ!$Q$5,0,0,"
                   f"MAX(MAX(期間データ!$M$5:$M${CSV_END})+MAX(期間データ!$O$5:$O${CSV_END}),1),1)"))

    ws.freeze_panes = "C14"

    # ======================================================== CSV貼付A/B =====
    for sheet_name, tab, label, csv_path in [
            ("CSV貼付A", AMBER, "期間A（直近金土日）", csv_a),
            ("CSV貼付B", AMBER2, "期間B（前週 金〜木）", csv_b)]:
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = tab
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.paperSize = 9
        ws.print_area = "A1:N45"        # 印刷用途は想定しないため先頭部のみ

        ws.column_dimensions["A"].width = 13
        for c_idx in range(2, NCOL + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 11
        ws.column_dimensions["N"].width = 34      # 商品名(全角16字まで見切れなし)
        ws.column_dimensions["J"].width = 41      # 作品名(元データが全角20字で切られている)
        ws.column_dimensions["L"].width = 35      # 支払先名(「本社　番組編成部　マーケティング室」等)
        # 商品コード(M)は13桁のためGeneralだと指数表記になる(幅も13桁ぶん確保)
        ws.column_dimensions["M"].number_format = "0"
        ws.column_dimensions["M"].width = 15
        # シート保護: 貼り付け領域(A〜AH列)のみ編集可。状態表示・見出しはロックし、
        # シート全体選択の貼り付け事故をブロックする
        for c_idx in range(1, NCOL + 1):
            ws.column_dimensions[get_column_letter(c_idx)].protection = UNLOCKED

        ws.row_dimensions[1].height = 34
        title_band(ws, f"A1:{get_column_letter(NCOL)}1",
                   f"　📋 {sheet_name}｜{label}の「売上・在庫・原価」CSV")
        ws.row_dimensions[2].height = 30
        note(ws, "A2:J2",
             "① CSVを開いて全選択→コピー（Ctrl+A → Ctrl+C）　"
             "② 下のオレンジのセル（A4）を選択　③ 右クリック→『値の貼り付け』。"
             "ヘッダー行ごと貼ってOKです（最大1000行）。貼り替えるときは、先に前回のデータ"
             "（5行目以降）を選択して削除してください。",
             9, GRAY, wrap=True)
        ws.row_dimensions[3].height = 18
        ws.merge_cells("A3:J3")
        ps, pe = parse_ymd("$D$5"), parse_ymd("$E$5")
        ws["A3"] = (f'="貼付行数: "&COUNTA($N$5:$N${CSV_END})&"行"&IF($N$5="","",'
                    f'"｜対象期間: "&IF(OR({ps}=-1,{pe}=-1),"—",'
                    f'TEXT({ps},"m/d")&"〜"&TEXT({pe},"m/d")))')
        style_range(ws, "A3:J3", font=fnt(9, True, "5B6472"), alignment=align("left"))

        ws.row_dimensions[4].height = 20
        for c_idx, h in enumerate(CSV_HEADERS, start=1):
            ref = f"{get_column_letter(c_idx)}4"
            ws[ref] = h
            style_range(ws, ref, font=fnt(8.5, True, "FFFFFF"), fl=fill(NAVY),
                        alignment=align("center"), border=BORDER_LIGHT)
        # 貼り付け開始セル(A4)を強調。値の貼り付けなら塗り・メモは残る
        style_range(ws, "A4", font=fnt(9, True, "7A4A00"), fl=fill("FFB84C"),
                    alignment=align("center"),
                    border=Border(left=coral_side, right=coral_side,
                                  top=coral_side, bottom=coral_side))
        c = mk_comment("👉 貼り付けはここから！\n"
                    "CSV全体をコピー（Ctrl+A→Ctrl+C）して、このセル（A4）を選択し、\n"
                    "右クリック→「値の貼り付け」。ヘッダー行ごと貼ってOKです。")
        ws["A4"].comment = c
        for c_idx in range(1, NCOL + 1):         # ヘッダー行ごと貼れるよう4行目も編集可
            ws.cell(row=4, column=c_idx).protection = UNLOCKED

        for i in range(350):                     # 目安の枠線(貼付は1000行まで有効)
            r = 5 + i
            for c_idx in range(1, NCOL + 1):
                cell = ws.cell(row=r, column=c_idx)
                cell.border = BORDER_HAIR
                cell.font = fnt(9)
                cell.protection = UNLOCKED       # セル書式が列書式に勝つため個別指定
                if c_idx == 13:
                    cell.number_format = "0"

        if csv_path:
            rows = read_csv_rows(csv_path)[:CSV_MAX]
            for i, row in enumerate(rows):
                r = 5 + i
                for c_idx, v in enumerate(row[:NCOL], start=1):
                    ws.cell(row=r, column=c_idx).value = v

        ws.freeze_panes = "A5"
        ws.protection.sheet = True

    # ============================================== 係数算出・係数貼付①〜④ ==
    # 時間帯係数の較正シート群(実測候補は準備数計算のK列に連動)
    from build_calib import add_calib_sheets
    add_calib_sheets(wb, mso_csvs, close=parse_clock(close), open_=parse_clock(open_))

    # ------------------------------------------------------------------ save -
    wb.properties.title = "コンセッション事前準備数ツール"
    wb.properties.creator = "TOHOシネマズ新宿 コンセッション"
    wb.save(out_path)
    print("saved:", out_path)


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="/home/user/claude-code-practice/concession-prep/TOHO新宿_コンセッション準備数ツール.xlsx")
    ap.add_argument("--csv-a")
    ap.add_argument("--csv-b")
    ap.add_argument("--select", choices=["A", "B", "AVG"], default="A")
    ap.add_argument("--att-a", help="期間Aの動員数3日分(カンマ区切り)")
    ap.add_argument("--att-b", help="期間Bの動員数7日分(カンマ区切り)")
    ap.add_argument("--peak", type=int)
    ap.add_argument("--preset", default="平常（基準）")
    for k in range(1, 5):
        ap.add_argument(f"--mso{k}", help=f"係数貼付{'①②③④'[k - 1]}に入れるMSO商品CSV(金曜1日分)")
    ap.add_argument("--close", default="22:00",
                    help="係数算出の閉店時刻(基本22:00。レイト日は 26:00 や 2:00=翌2時も可)")
    ap.add_argument("--open", dest="open_", default="08:00",
                    help="係数算出の開店時刻(基本08:00。早朝上映のある日は 06:00 等)")
    ap.add_argument("--show-notes", metavar="XLSX",
                    help="既存xlsxの仕上げ: フォント名の書き戻し+貼り付けメモの常時表示化(再計算後に実行)")
    a = ap.parse_args()
    if a.show_notes:
        force_font(a.show_notes)
        n = show_paste_comments(a.show_notes)
        print(f"finalized: font={FONT_NAME}, notes patched: {n} vml file(s) in {a.show_notes}")
        raise SystemExit(0)
    build(a.out, csv_a=a.csv_a, csv_b=a.csv_b, select=a.select,
          att_a=[int(x) for x in a.att_a.split(",")] if a.att_a else None,
          att_b=[int(x) for x in a.att_b.split(",")] if a.att_b else None,
          peak=a.peak, preset=a.preset,
          mso_csvs=(a.mso1, a.mso2, a.mso3, a.mso4), close=a.close, open_=a.open_)
