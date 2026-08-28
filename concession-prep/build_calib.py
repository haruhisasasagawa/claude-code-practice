# -*- coding: utf-8 -*-
"""
時間帯係数の較正シート群（係数算出＋係数貼付①〜④）を本体ワークブックへ追加するモジュール

MSO商品CSV(注文明細・28列): K=売店売上日付, M=注文時間, S=ステータス, T=取消区分,
Z=販売数量, AB=商品区分。金曜1日分で出力したCSVを係数貼付①〜④にそのまま貼ると、
時間帯(5段階)ごとの販売ペースから時間帯係数の候補を自動算出し、
「準備数計算」プリセット表の実測候補列(K列)にも表示される。

build_tool.py の build() から add_calib_sheets(wb, csvs) として呼ばれる。
単体では実行しない(本体の生成は python build_tool.py)。
"""
import csv

from openpyxl.formatting.rule import DataBarRule, FormulaRule
from openpyxl.styles import Border, Font, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

from build_tool import (BORDER_HAIR, BORDER_INPUT, BORDER_LIGHT, CHIP_AMBER,
                        CHIP_CORAL, CHIP_NAVY, CHIP_TEAL, CORAL, F_INPUT, F_ZEBRA,
                        FONT_NAME, GRAY, GREEN, INK, JUDGE_FEW, JUDGE_NODATA,
                        JUDGE_NONE, JUDGE_USE, N_SLOTS, NAVY, RED, ROW_P0, TEAL,
                        WAVE_ROW0, WAVE_SHEET, WEEKDAY_JA, align, chip, coral_side,
                        fill, fnt, mk_comment, note, style_range, title_band)

MSO_HEADERS = ["操作モード", "伝票番号", "伝票枝番", "サイトコード", "サイト名",
               "劇場コード", "劇場名", "販売場所コード", "販売場所名", "端末番号",
               "売店売上日付", "注文日付", "注文時間", "受付時間", "提供時間", "提供済時間",
               "注文金額", "注文番号", "ステータス", "取消区分", "会員番号", "商品明細番号",
               "商品コード", "商品名", "販売単価", "販売数量", "販売金額", "商品区分"]
MSO_NCOL = len(MSO_HEADERS)      # 28列: K=日付(11), M=時間(13), S=状態(19), T=取消(20),
#                                        Z=数量(26), AB=区分(28)
MSO_MAX = 30000                  # 貼付データ最大行数(金曜1日分を想定。5〜30004行目)
MSO_END = 4 + MSO_MAX
CALIB_SHEETS = ["係数貼付①", "係数貼付②", "係数貼付③", "係数貼付④"]
BANDS = ["① 朝一", "② 昼ピーク", "③ 夕方", "④ 夜ピーク", "⑤ レイト"]

PURPLE = "7C5CB0"                # 係数算出タブ
PURPLE_L = "B9A3D6"              # 係数貼付タブ
UNLOCKED = Protection(locked=False)


def read_mso_rows(path):
    raw = open(path, "rb").read()
    for enc in ("cp932", "utf-8-sig", "utf-8"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    rows = [r for r in csv.reader(text.splitlines()) if len(r) == MSO_NCOL]
    out = []
    for r in rows[1:]:
        conv = []
        for v in r:
            v = v.strip()
            if v == "":
                conv.append(None)
                continue
            try:
                f = float(v)
                conv.append(int(f) if f == int(f) else f)
            except ValueError:
                conv.append(v)
        out.append(conv)
    return out


def mso_date_expr(x, fallback):
    """売店売上日付セルを日付シリアルに解釈する式。数値セル・"2026/07/17"・
    "2026/7/10"(ゼロ埋めなし。実データに存在する表記ゆれ)のいずれにも対応し、
    解釈不能なら fallback を返す(DATEVALUEはロケール依存のため最後の砦のみ)"""
    rest = f"MID({x},6,9)"
    return (f'IF(ISNUMBER({x}),INT({x}),'
            f'IFERROR(DATE(VALUE(LEFT({x},4)),'
            f'VALUE(LEFT({rest},FIND("/",{rest})-1)),'
            f'VALUE(MID({rest},FIND("/",{rest})+1,9))),'
            f'IFERROR(DATEVALUE({x}),{fallback})))')


def add_calib_sheets(wb, csvs=(None, None, None, None), close=22 / 24, open_=8 / 24):
    """本体ワークブックへ「係数算出」「商品別の波」「係数貼付①〜④」を追加する。
    close: 閉店時刻の初期値(時/24。22:00なら22/24、26:00なら26/24)
    open_: 開店時刻の初期値(時/24。早朝上映のある劇場は6:00等)"""

    # ========================================================== 係数算出 =====
    ws = wb.create_sheet("係数算出")
    ws.sheet_properties.tabColor = PURPLE
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = 9
    ws.print_area = "A1:L26"

    for c, w in {"A": 2.5, "B": 15, "C": 13, "D": 10, "E": 10, "F": 10, "G": 10,
                 "H": 11, "I": 10, "J": 11, "K": 11, "L": 13}.items():
        ws.column_dimensions[c].width = w
    for c in "MNO":                       # ⚙内部ヘルパー列は隠す
        ws.column_dimensions[c].hidden = True

    ws.row_dimensions[1].height = 34
    title_band(ws, "B1:L1", "　⏱ 係数算出｜金曜4週分の時間帯の波（月1回・任意）")
    ws.row_dimensions[2].height = 18
    note(ws, "B2:L2", "係数候補 ＝ 帯の1時間あたり販売個数 ÷ 全帯平均の1時間あたり販売個数（セット親・取消・払戻は除外）", 9)

    # 時間の区切り(編集OK)
    ws.row_dimensions[3].height = 20
    chip(ws, "B3", "  ⏰ 時間の区切り", CHIP_AMBER, INK, 9)
    tb = [("C3", "開店", open_), ("D3", "朝→昼", 11 / 24), ("E3", "昼→夕", 15 / 24),
          ("F3", "夕→夜", 18 / 24), ("G3", "夜→レイト", 21 / 24), ("H3", "閉店", close)]
    ws.row_dimensions[4].height = 18
    for ref, label, val in tb:
        note(ws, ref, label, 8, GRAY, h="center")
        vref = ref[0] + "4"
        ws[vref] = val
        style_range(ws, vref, font=fnt(9, True), fl=fill(F_INPUT),
                    alignment=align("center"), border=BORDER_INPUT, num="h:mm")
        ws[vref].protection = UNLOCKED
    ws["H3"].comment = mk_comment("閉店時刻。基本は 22:00。レイト営業日は 26:00 または 翌2:00 の"
                               "形で入力してください(24時間超え表記対応)。同日閉店(21時以降)と"
                               "翌日閉店(24:00〜)は自動判別します。")
    # 閉店の正規化(印刷範囲外のN・O列): O4=同日側の上限、N4=翌日側の締めフラクション
    #   26:00/翌2:00 → O4=24:00・N4=2:00 / 22:00 → O4=22:00・N4=0 /
    #   開店とレイト開始の間(20:00等の設定ミス) → 空の帯になりB23に⚠
    # 翌側の締めは開店時刻でクランプする(33:00等の打ち間違いで帯⑤の翌日側窓が
    # 帯①と重なり二重計上になるのを構造的に防ぐ。B23に専用⚠も出す)
    ws["N4"] = ('=IF($H$4>=1,MIN(MOD($H$4,1),$C$4),'
                'IF($H$4>=$G$4,0,IF($H$4<=$C$4,$H$4,0)))')
    ws["O4"] = ('=IF($H$4>=1,1,'
                'IF($H$4>=$G$4,$H$4,IF($H$4<=$C$4,1,$H$4)))')
    for ref in ("N4", "O4"):
        style_range(ws, ref, font=fnt(8.5, False, GRAY), alignment=align("center"), num="h:mm")

    # 帯ごとの窓: ①[開店,朝→昼) ②[朝→昼,昼→夕) ③[昼→夕,夕→夜) ④[夕→夜,夜→レイト)
    #             ⑤[夜→レイト,24:00)+[0:00,閉店)
    ws.row_dimensions[6].height = 26
    heads = [("B6", "時間帯"), ("C6", "時間の目安"), ("D6", "1週目"), ("E6", "2週目"),
             ("F6", "3週目"), ("G6", "4週目"), ("H6", "平均個数"), ("I6", "帯の長さ"),
             ("J6", "1h当り"), ("K6", "係数候補"), ("L6", "転記用(丸め)")]
    for ref, text in heads:
        style_range(ws, ref, font=fnt(9, True, "FFFFFF"), fl=fill(NAVY),
                    alignment=align("center", "center", True), border=BORDER_LIGHT)
        ws[ref] = text
    style_range(ws, "K6:L6", font=fnt(9, True, "FFFFFF"), fl=fill(CORAL),
                alignment=align("center", "center", True), border=BORDER_LIGHT)
    ws["K6"] = "係数候補"
    ws["L6"] = "転記用(丸め)"

    starts = ["$C$4", "$D$4", "$E$4", "$F$4", "$G$4"]
    ends = ["$D$4", "$E$4", "$F$4", "$G$4", None]      # ⑤は21時〜24時+0時〜閉店
    for bi, name in enumerate(BANDS):
        r = 7 + bi
        ws.row_dimensions[r].height = 20
        ws[f"B{r}"] = name
        ws[f"C{r}"] = (f'=TEXT({starts[bi]},"h:mm")&"〜"&' +
                       (f'TEXT({ends[bi]},"h:mm")' if ends[bi]
                        else 'TEXT($H$4,"h:mm")&IF(OR($H$4>=1,$H$4<=$C$4),"(翌)","")'))
        for wi, sheet in enumerate(CALIB_SHEETS):
            col = get_column_letter(4 + wi)
            rng_t = f"{sheet}!$AE$5:$AE${MSO_END}"
            rng_q = f"{sheet}!$AF$5:$AF${MSO_END}"
            # 貼付シートの正常フラグ(AD3)が0の週(別CSVの誤貼付・式の破損)は集計しない
            ok = f"{sheet}!$AD$3=1"
            # ⑤レイト帯は正規化済みの閉店(O4=同日上限・N4=翌日側締め)で区切る。
            # 22:00(同日)と26:00=翌2:00(翌日)のどちらの閉店にも同じ式で対応
            if ends[bi]:
                ws[f"{col}{r}"] = (f'=IF({ok},SUMIFS({rng_q},{rng_t},">="&{starts[bi]},'
                                   f'{rng_t},"<"&{ends[bi]}),0)')
            else:
                ws[f"{col}{r}"] = (f'=IF({ok},SUMIFS({rng_q},{rng_t},">="&{starts[bi]},'
                                   f'{rng_t},"<"&$O$4)'
                                   f'+SUMIFS({rng_q},{rng_t},"<"&$N$4),0)')
        ws[f"H{r}"] = f'=IF($H$13=0,"",SUM(D{r}:G{r})/$H$13)'
        if ends[bi]:
            ws[f"I{r}"] = f'=({ends[bi]}-{starts[bi]})*24'
        else:
            ws[f"I{r}"] = f'=($O$4-{starts[bi]})*24+$N$4*24'
        ws[f"J{r}"] = f'=IF(OR($H{r}="",$I{r}<=0),"",$H{r}/$I{r})'
        ws[f"K{r}"] = f'=IF(OR($J{r}="",$J$12=""),"",IF($J$12<=0,"",$J{r}/$J$12))'
        ws[f"L{r}"] = f'=IF($K{r}="","—",ROUND($K{r}/0.05,0)*0.05)'
        style_range(ws, f"B{r}", font=fnt(9.5, True), alignment=align("left"))
        style_range(ws, f"C{r}", font=fnt(8.5, False, GRAY), alignment=align("center"))
        style_range(ws, f"D{r}:G{r}", font=fnt(9.5, False, "5B6472"),
                    alignment=align("center"), num="#,##0")
        style_range(ws, f"H{r}", font=fnt(9.5, False, "5B6472"), alignment=align("center"), num="#,##0.0")
        style_range(ws, f"I{r}", font=fnt(9.5, False, GRAY), alignment=align("center"), num='0.0"h"')
        style_range(ws, f"J{r}", font=fnt(9.5, False, "5B6472"), alignment=align("center"), num="#,##0.0")
        style_range(ws, f"K{r}", font=fnt(10.5, True, CORAL), fl=fill("FFF1EE"),
                    alignment=align("center"), num="0.00")
        style_range(ws, f"L{r}", font=fnt(10.5, True, CORAL), fl=fill("FFF1EE"),
                    alignment=align("center"), num='0.00"倍"')
        if bi % 2:
            for col in "BCDEFGHIJ":
                ws[f"{col}{r}"].fill = fill(F_ZEBRA)
        for col in "BCDEFGHIJ":
            ws[f"{col}{r}"].border = Border(bottom=BORDER_HAIR.bottom, left=BORDER_HAIR.left,
                                            right=BORDER_HAIR.right)
        # 最下行(⑤レイト)は下辺もコーラルにして強調枠を閉じる
        k_bottom = coral_side if bi == len(BANDS) - 1 else BORDER_HAIR.bottom
        ws[f"K{r}"].border = Border(left=coral_side, right=BORDER_HAIR.right, bottom=k_bottom)
        ws[f"L{r}"].border = Border(right=coral_side, bottom=k_bottom)

    # 合計行・全日ペース
    ws.row_dimensions[12].height = 20
    style_range(ws, "B12", font=fnt(9.5, True), alignment=align("left"))
    ws["B12"] = "帯内合計 / 全日ペース"
    for wi in range(4):
        col = get_column_letter(4 + wi)
        ws[f"{col}12"] = f"=SUM({col}7:{col}11)"
        style_range(ws, f"{col}12", font=fnt(9.5, True, "5B6472"), alignment=align("center"), num="#,##0")
    ws["H12"] = '=IF($H$13=0,"",SUM(H7:H11))'
    ws["I12"] = "=SUM(I7:I11)"
    ws["J12"] = '=IF(OR($H$12="",$I$12<=0),"",$H$12/$I$12)'
    style_range(ws, "H12", font=fnt(9.5, True, "5B6472"), alignment=align("center"), num="#,##0.0")
    style_range(ws, "I12", font=fnt(9.5, True, GRAY), alignment=align("center"), num='0.0"h"')
    style_range(ws, "J12", font=fnt(9.5, True, "5B6472"), alignment=align("center"), num="#,##0.0")
    style_range(ws, "B12:J12", border=Border(top=BORDER_LIGHT.top))
    style_range(ws, "K12:L12", border=Border(top=coral_side))

    # 使った週数
    ws.row_dimensions[13].height = 16
    # ラベルは値(H13)の左隣まで伸ばして右寄せ(離れているとどの値のラベルか読めない)
    note(ws, "B13:G13", "集計に使った週数", 8.5, GRAY, h="right")
    ws["H13"] = "=SUMPRODUCT((D12:G12>0)*1)"
    style_range(ws, "H13", font=fnt(9, True, TEAL), alignment=align("center"), num='0"週"')

    # 週ごとの状態表示
    ws.row_dimensions[15].height = 18
    chip(ws, "B15:L15", "  各週の貼付状況", CHIP_NAVY, NAVY, 9)
    for wi, sheet in enumerate(CALIB_SHEETS):
        r = 16 + wi
        ws.row_dimensions[r].height = 16
        note(ws, f"B{r}", f"  {wi + 1}週目（{sheet}）", 8.5, GRAY)
        ws.merge_cells(f"C{r}:L{r}")
        ws[f"C{r}"] = f"={sheet}!$A$3"
        style_range(ws, f"C{r}:L{r}", font=fnt(8.5, False, "5B6472"), alignment=align("left"))

    # 転記案内
    ws.row_dimensions[21].height = 22
    chip(ws, "B21:L21", "  📋 転記のしかた", CHIP_CORAL, INK, 9.5)
    ws.row_dimensions[22].height = 30
    note(ws, "B22:L22",
         "上の表の「転記用(丸め)」①〜⑤の5つの値を、「準備数計算」シート右上の時間帯プリセット表"
         "（J4:J8）に手で入力してください（プリセット表の右の「実測候補」列にも同じ値が自動表示"
         "されます。採用は手入力で）。平常（基準）は 1.0倍 のままでOKです。",
         9, INK, wrap=True)
    ws.row_dimensions[23].height = 16
    bad_any = "+".join(f"({s}!$AD$3=0)" for s in CALIB_SHEETS)
    # N5=区切りの外の販売個数(有効週の対象個数合計−帯内合計)、
    # O5=同日閉店疑いの週数(翌日閉店設定なのに22時以降の販売が0個の有効週)
    valid_total = "(" + "+".join(
        f"IF({s}!$AD$3=1,SUM({s}!$AF$5:$AF${MSO_END}),0)" for s in CALIB_SHEETS) + ")"
    ws["N5"] = f"={valid_total}-SUM($D$12:$G$12)"
    susp_terms = []
    for wi, s in enumerate(CALIB_SHEETS):
        wcol = "DEFG"[wi]
        late = (f'SUMIFS({s}!$AF$5:$AF${MSO_END},{s}!$AE$5:$AE${MSO_END},">="&TIME(22,0,0))'
                f'+SUMIFS({s}!$AF$5:$AF${MSO_END},{s}!$AE$5:$AE${MSO_END},"<"&$N$4)')
        susp_terms.append(f'IF(AND(${wcol}$12>0,{late}=0),1,0)')
    ws["O5"] = f'=IF($N$4=0,0,{"+".join(susp_terms)})'
    for ref in ("N5", "O5"):
        style_range(ws, ref, font=fnt(8.5, False, GRAY), alignment=align("center"), num="#,##0")
    ws["B23"] = ('=TRIM('
                 'IF($H$13=0,"⚠ まだ有効なデータが貼られていません。係数貼付①〜④に金曜1日分の'
                 'MSO商品CSVを貼ってください（貼らなくても本体はプリセットの既定係数で使えます）。","")&" "&'
                 f'IF(({bad_any})>0,'
                 '"⚠ 貼付シートに問題があり集計から除外した週があります（上の貼付状況を確認）。","")&" "&'
                 f'IF(AND($H$4>$C$4,$H$4<$G$4),'
                 '"⚠ 閉店時刻がレイト開始（夜→レイトの区切り）より前になっています。'
                 '時間の区切りを確認してください。","")&" "&'
                 'IF(AND($H$4>=1,MOD($H$4,1)>$C$4),'
                 '"⚠ 閉店（翌）が開店時刻を超えています。閉店の入力を確認してください'
                 '（例: 26:00＝翌2時。翌側は開店までで打ち切って集計中）。","")&" "&'
                 'IF($N$5>0,'
                 '"⚠ 時間の区切りの外の販売が "&TEXT($N$5,"#,##0")&" 個あります'
                 '（開店・閉店の時刻を確認。この分は係数の集計対象外です）。","")&" "&'
                 'IF($O$5>0,'
                 '"⚠ 22時以降の販売が0個の週が"&$O$5&"週あります（レイト営業の無い金曜（同日閉店）'
                 'が混ざっている可能性。閉店の設定と貼付データを確認してください）。",""))')
    ws.merge_cells("B23:L23")
    style_range(ws, "B23:L23", font=fnt(8.5, True, "D14343"), alignment=align("left"))
    # 時間の区切り(C4:H4)以外は自動計算のためシート保護(パスワード無し)
    ws.protection.sheet = True

    # ======================================================== 商品別の波 =====
    ws = wb.create_sheet(WAVE_SHEET)
    ws.sheet_properties.tabColor = PURPLE
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.paperSize = 9
    ws.print_area = "A1:N30"

    ws.column_dimensions["A"].width = 2.5
    ws.column_dimensions["B"].width = 30
    for c in "CDEFG":
        ws.column_dimensions[c].width = 8.5
    ws.column_dimensions["H"].width = 9
    for c in "IJKLM":
        ws.column_dimensions[c].width = 8.5
    ws.column_dimensions["N"].width = 22
    ws.column_dimensions["O"].width = 2.5
    for c in "PQRST":
        ws.column_dimensions[c].width = 8
        ws.column_dimensions[c].hidden = True

    ws.row_dimensions[1].height = 34
    title_band(ws, "B1:N1", "　⏱ 商品別の波｜商品ごとの時間帯パターン（自動）")
    ws.row_dimensions[2].height = 18
    note(ws, "B2:N2",
         "商品＝期間データの登録商品（B14〜B33）。個数＝係数貼付①〜④（有効な週）の合計。"
         "構成比で1日の売れ方の推移が見え、商品別係数は準備数計算「⑤ 商品別の波」で作る数に適用できます。", 9)

    ws.row_dimensions[3].height = 20
    chip(ws, "B3:C3", "  ⚙ 係数を使う最低個数", CHIP_AMBER, INK, 9)
    ws["D3"] = 30
    style_range(ws, "D3", font=fnt(9, True), fl=fill(F_INPUT),
                alignment=align("center"), border=BORDER_INPUT, num='#,##0"個"')
    ws["D3"].protection = UNLOCKED
    note(ws, "E3:N3",
         "← 係数貼付①〜④の合計個数がこの数に満たない商品は、精度が低いため全体の時間帯係数で計算します（編集可）。", 8.5)
    dv_thr = DataValidation(type="whole", operator="between", formula1="1", formula2="999999",
                            showErrorMessage=True)
    dv_thr.error = "最低個数は 1〜999,999 の整数で入力してください"
    dv_thr.errorTitle = "最低個数"
    ws.add_data_validation(dv_thr)
    dv_thr.add("D3")

    ws.row_dimensions[5].height = 18
    chip(ws, "C5:G5", " 構成比（1日の売れ方の推移）", CHIP_TEAL, INK, 8.5)
    chip(ws, "I5:M5", " 商品別係数（0.05刻み・自動）", CHIP_CORAL, INK, 8.5)
    ws.row_dimensions[6].height = 26
    style_range(ws, "B6", font=fnt(9, True, "FFFFFF"), fl=fill(NAVY),
                alignment=align("center", "center", True), border=BORDER_LIGHT)
    ws["B6"] = "商品名"
    for bi, name in enumerate(BANDS):
        for base in ("C", "I"):
            ref = f"{get_column_letter(ord(base) - 64 + bi)}6"
            ws[ref] = name
            style_range(ws, ref, font=fnt(8, True, "FFFFFF"),
                        fl=fill(CORAL if base == "I" else NAVY),
                        alignment=align("center", "center", True), border=BORDER_LIGHT)
    for ref, text in [("H6", "合計個数"), ("N6", "判定")]:
        ws[ref] = text
        style_range(ws, ref, font=fnt(9, True, "FFFFFF"), fl=fill(NAVY),
                    alignment=align("center", "center", True), border=BORDER_LIGHT)
    # 非表示のP〜T列に帯別個数を置く(印刷ににじむためラベルは置かない)
    starts5 = ["$C$4", "$D$4", "$E$4", "$F$4", "$G$4"]
    ends5 = ["$D$4", "$E$4", "$F$4", "$G$4", None]
    thr = 'IF(ISNUMBER($D$3),$D$3,30)'
    for i in range(N_SLOTS):
        wr = WAVE_ROW0 + i
        pr = ROW_P0 + i
        ws.row_dimensions[wr].height = 19
        ws[f"B{wr}"] = f'=IF(期間データ!B{pr}="","",期間データ!B{pr})'
        # 帯別個数(非表示P〜T列)。無効な週(AD3=0)は除外。⑤は閉店(翌)をMODで正規化
        for bi in range(5):
            col = get_column_letter(16 + bi)
            terms = []
            for sheet in CALIB_SHEETS:
                q = f"{sheet}!$AF$5:$AF${MSO_END}"
                t = f"{sheet}!$AE$5:$AE${MSO_END}"
                x = f"{sheet}!$X$5:$X${MSO_END}"
                if ends5[bi]:
                    s = (f'IF({sheet}!$AD$3=1,SUMIFS({q},{x},$B{wr},'
                         f'{t},">="&係数算出!{starts5[bi]},{t},"<"&係数算出!{ends5[bi]}),0)')
                else:
                    s = (f'IF({sheet}!$AD$3=1,SUMIFS({q},{x},$B{wr},'
                         f'{t},">="&係数算出!$G$4,{t},"<"&係数算出!$O$4)'
                         f'+SUMIFS({q},{x},$B{wr},{t},"<"&係数算出!$N$4),0)')
                terms.append(s)
            ws[f"{col}{wr}"] = f'=IF($B{wr}="","",{"+".join(terms)})'
            style_range(ws, f"{col}{wr}", font=fnt(8.5, False, GRAY), alignment=align("center"))
        ws[f"H{wr}"] = f'=IF($B{wr}="","",SUM($P{wr}:$T{wr}))'
        ws[f"N{wr}"] = (f'=IF($B{wr}="","",IF(係数算出!$H$13=0,"{JUDGE_NODATA}",'
                        f'IF($H{wr}=0,"{JUDGE_NONE}",'
                        f'IF($H{wr}<{thr},"{JUDGE_FEW}","{JUDGE_USE}"))))')
        for bi in range(5):
            ccol = get_column_letter(3 + bi)
            pcol = get_column_letter(16 + bi)
            ws[f"{ccol}{wr}"] = f'=IF($B{wr}="","",IF($H{wr}=0,"—",{pcol}{wr}/$H{wr}))'
            kcol = get_column_letter(9 + bi)
            # 帯長が0以下(時間の区切りの逆転・同値)と帯個数0のときは"—"で全体係数へ
            # フォールバック(負の帯長は0.00が数値として適用される事故、帯個数0は
            # 「作らない→売れない→係数0→作らない」の自己成就と無警告の作る数0を防ぐ)
            ws[f"{kcol}{wr}"] = (f'=IF($B{wr}="","",IF($N{wr}<>"{JUDGE_USE}","—",'
                                 f'IF(OR(係数算出!$I${7 + bi}<=0,係数算出!$I$12<=0,{pcol}{wr}=0),"—",'
                                 f'IFERROR(ROUND(({pcol}{wr}/$H{wr})*'
                                 f'(係数算出!$I$12/係数算出!$I${7 + bi})/0.05,0)*0.05,"—"))))')
        style_range(ws, f"B{wr}", font=fnt(9.5), alignment=align("left"))
        style_range(ws, f"C{wr}:G{wr}", font=fnt(9, False, "5B6472"),
                    alignment=align("center"), num="0%")
        style_range(ws, f"H{wr}", font=fnt(9, True, "5B6472"), alignment=align("center"), num="#,##0")
        style_range(ws, f"I{wr}:M{wr}", font=fnt(9.5, True, CORAL), fl=fill("FFF1EE"),
                    alignment=align("center"), num='0.00"倍"')
        style_range(ws, f"N{wr}", font=fnt(8.5, False, GRAY), alignment=align("center"))
        if i % 2:
            for c in "BCDEFGHN":
                ws[f"{c}{wr}"].fill = fill(F_ZEBRA)
        for c in "BCDEFGHIJKLMN":
            ws[f"{c}{wr}"].border = BORDER_HAIR
    last_w = WAVE_ROW0 + N_SLOTS - 1
    bar = DataBarRule(start_type="num", start_value=0, end_type="max", color=TEAL, showValue=True)
    ws.conditional_formatting.add(f"C{WAVE_ROW0}:G{last_w}", bar)
    j_ok = FormulaRule(formula=[f'LEFT($N{WAVE_ROW0},1)="✔"'],
                       font=Font(name=FONT_NAME, size=8.5, color=GREEN))
    j_ng = FormulaRule(formula=[f'ISNUMBER(SEARCH("⚠",$N{WAVE_ROW0}))'],
                       font=Font(name=FONT_NAME, size=8.5, bold=True, color=RED))
    ws.conditional_formatting.add(f"N{WAVE_ROW0}:N{last_w}", j_ok)
    ws.conditional_formatting.add(f"N{WAVE_ROW0}:N{last_w}", j_ng)

    ws.row_dimensions[last_w + 2].height = 16
    note(ws, f"B{last_w + 2}:N{last_w + 2}",
         "※ 商品別係数 ＝ その商品の帯の1時間あたり個数 ÷ その商品の1日平均ペース（係数算出シートの区切りに連動）。"
         "個数0の帯は「—」＝全体の時間帯係数で計算します（作る数が無警告で0になるのを防ぐため）。", 8.5)
    ws.row_dimensions[last_w + 3].height = 16
    note(ws, f"B{last_w + 3}:N{last_w + 3}",
         "※ 「該当なし」はMSO側に同名の商品が無い場合です（商品名の表記が売上CSVと異なる可能性。その商品は全体の時間帯係数で計算されます）。", 8.5)
    ws.row_dimensions[last_w + 4].height = 16
    note(ws, f"B{last_w + 4}:N{last_w + 4}",
         "※ このシートは全て自動計算です（最低個数D3のみ編集可・シート保護済み）。", 8.5)
    ws.freeze_panes = "A7"
    # しきい値(D3)以外は自動計算のためシート保護(行対応が崩れる並べ替え等も防止)
    ws.protection.sheet = True

    # ======================================================= 係数貼付①〜④ ===
    for si, sheet_name in enumerate(CALIB_SHEETS):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = PURPLE_L
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 1
        ws.page_setup.paperSize = 9
        ws.print_area = "A1:N40"

        ws.column_dimensions["A"].width = 11
        for c_idx in range(2, MSO_NCOL + 1):
            ws.column_dimensions[get_column_letter(c_idx)].width = 11
        ws.column_dimensions["X"].width = 34          # 商品名(全角16字まで見切れなし)
        ws.column_dimensions["G"].width = 21          # 劇場名(ＴＯＨＯシネマズ新宿)
        ws.column_dimensions["I"].width = 17          # 販売場所名(モバイルオーダー)
        # 伝票番号(B)・商品コード(W)は12桁超のためGeneralだと指数表記になる。
        # 14桁が収まるよう幅も広げる(幅不足だと###表示)
        for c in ("B", "W"):
            ws.column_dimensions[c].number_format = "0"
            ws.column_dimensions[c].width = 16
        # シート保護: 貼り付け領域(A〜AB列)と対象日S3のみ編集可。AE/AF式・状態表示・
        # 正常フラグはロックし、シート全体選択の貼り付けで式が消える事故をブロックする
        for c_idx in range(1, MSO_NCOL + 1):
            ws.column_dimensions[get_column_letter(c_idx)].protection = UNLOCKED
        ws.column_dimensions["AC"].width = 2.5        # 緩衝
        ws.column_dimensions["AD"].width = 10
        ws.column_dimensions["AE"].width = 9
        ws.column_dimensions["AF"].width = 9

        ws.row_dimensions[1].height = 34
        title_band(ws, f"A1:{get_column_letter(MSO_NCOL)}1",
                   f"　⏱ {sheet_name}｜{si + 1}週目の金曜のMSO商品CSV（時間帯係数の較正用）")
        ws.row_dimensions[2].height = 30
        # 折り返し位置は自前の改行で制御(自動折返しだと閉じ括弧が行頭に落ちる)
        note(ws, "A2:N2",
             "① 金曜1日分で出力したMSO商品CSVを開いて全選択→コピー（Ctrl+A → Ctrl+C）　"
             "② 下のオレンジのセル（A4）を選択　③ 右クリック→『値の貼り付け』\n"
             f"ヘッダー行ごと貼ってOKです（最大{MSO_MAX:,}行）。貼り替えるときは、前回のデータ"
             "（5行目以降のA〜AB列）だけを選択してDeleteで消してください（行ごと削除しないこと）。",
             9, GRAY, wrap=True)

        ws.row_dimensions[3].height = 18
        ws.merge_cells("A3:P3")
        # AD3=正常フラグ: ヘッダー行の照合(別CSVの誤貼付・貼付位置ズレ検知)と
        # ヘルパー式AE/AFの生存数チェック(行削除による式の破損検知)。0の週は係数算出で無視される
        ws["AD3"] = (f'=IF(OR($K$4<>"売店売上日付",COUNTA($AE$5:$AE${MSO_END})<{MSO_MAX},'
                     f'COUNT($AF$5:$AF${MSO_END})<{MSO_MAX}),0,1)')
        style_range(ws, "AD3", font=fnt(8.5, False, GRAY), alignment=align("center"))
        ws["A3"] = (f'=IF($AD$3=0,'
                    f'IF($K$4<>"売店売上日付",'
                    f'"⚠ 貼り付け内容がMSO商品CSVではないか、位置がずれています'
                    f'（ヘッダーごとならA4、データのみならA5から。このシートは集計から除外中）",'
                    f'"⚠ 内部の集計式（AE・AF列）が一部消えています（行ごと削除が原因。'
                    f'配布元ファイルのシートから作り直してください。このシートは集計から除外中）"),'
                    f'IF($M$5="","（未貼付）",'
                    f'"貼付 "&COUNTA($M$5:$M${MSO_END})&"行"&'
                    f'IF($AD$2="","｜⚠ 日付を読み取れません",'
                    f'"｜対象日: "&TEXT($AD$2,"m/d")&"（"&CHOOSE(WEEKDAY($AD$2),{WEEKDAY_JA})&"）'
                    f'｜対象個数 "&TEXT(SUM($AF$5:$AF${MSO_END}),"#,##0")&"個"&'
                    f'IF(WEEKDAY($AD$2)<>6,"｜※金曜ではありません","")&'
                    f'IF(SUMPRODUCT(($K$5:$K${MSO_END}<>"")*($K$5:$K${MSO_END}<>$K$5))>0,'
                    f'"｜⚠ 複数日が混在（対象日以外は無視）","")&'
                    f'IF($M${MSO_END + 1}<>"","｜⚠ {MSO_MAX:,}行を超えています（超過分は集計対象外）","")'
                    f')))')
        style_range(ws, "A3:P3", font=fnt(9, True, "5B6472"), alignment=align("left"))
        note(ws, "Q3:R3", "対象日(空欄=自動):", 8, GRAY, h="right")
        style_range(ws, "S3", font=fnt(9, True), fl=fill(F_INPUT),
                    alignment=align("center"), border=BORDER_INPUT, num="m/d")
        ws["S3"].protection = UNLOCKED
        # 2000〜2100年の範囲外(誤貼付で数値コード等が来た場合)は空欄に落とす。
        # AD2はm/d日付書式のため、範囲外シリアルを返すとLibreOfficeが#VALUE!で書き出す
        ad2_inner = mso_date_expr("$K$5", "-1")
        ws["AD2"] = (f'=IF($S$3<>"",$S$3,IF($K$5="","",'
                     f'IFERROR(IF(AND({ad2_inner}>=36526,{ad2_inner}<=73415),{ad2_inner},""),"")))')
        note(ws, "AD1", "⚙対象日", 8, GRAY)
        style_range(ws, "AD2", font=fnt(8.5, False, GRAY), alignment=align("center"), num="m/d")

        ws.row_dimensions[4].height = 20
        for c_idx, h in enumerate(MSO_HEADERS, start=1):
            ref = f"{get_column_letter(c_idx)}4"
            ws[ref] = h
            style_range(ws, ref, font=fnt(8.5, True, "FFFFFF"), fl=fill(NAVY),
                        alignment=align("center"), border=BORDER_LIGHT)
        style_range(ws, "A4", font=fnt(9, True, "7A4A00"), fl=fill("FFB84C"),
                    alignment=align("center"),
                    border=Border(left=coral_side, right=coral_side,
                                  top=coral_side, bottom=coral_side))
        ws["A4"].comment = mk_comment("👉 貼り付けはここから！\n"
                                   "金曜1日分のMSO商品CSVを全選択コピーして、このセル（A4）を選択し、\n"
                                   "右クリック→「値の貼り付け」。ヘッダー行ごと貼ってOKです。")
        for c_idx in range(1, MSO_NCOL + 1):      # ヘッダー行ごと貼れるよう4行目も編集可
            ws.cell(row=4, column=c_idx).protection = UNLOCKED
        for ref, text in [("AE4", "⚙時刻値"), ("AF4", "⚙対象個数")]:
            note(ws, ref, text, 8, GRAY, h="center")

        # ヘルパー列: AE=時刻値, AF=集計対象個数(日付・状態・区分・数量のフィルタ込み)
        for i in range(MSO_MAX):
            rr = 5 + i
            ws[f"AE{rr}"] = (f'=IF($M{rr}="","",IFERROR(IF(ISNUMBER($M{rr}),MOD($M{rr},1),'
                             f'TIMEVALUE($M{rr})),""))')
            ws[f"AF{rr}"] = (f'=IF(OR($M{rr}="",$AE{rr}=""),0,'
                             f'IF({mso_date_expr(f"$K{rr}", "-1")}<>$AD$2,0,'
                             f'IF(OR($AB{rr}="セット親",$S{rr}<>"提供済",$T{rr}<>"販売"),0,'
                             f'IF(ISNUMBER($Z{rr}),MAX(0,$Z{rr}),0))))')

        for i in range(200):                      # 目安の枠線
            rr = 5 + i
            for c_idx in range(1, MSO_NCOL + 1):
                cell = ws.cell(row=rr, column=c_idx)
                cell.border = BORDER_HAIR
                cell.font = fnt(9)
                cell.protection = UNLOCKED        # セル書式が列書式に勝つため個別指定
                if c_idx in (2, 23):
                    cell.number_format = "0"

        rows = read_mso_rows(csvs[si]) if csvs[si] else []
        for i, row in enumerate(rows[:MSO_MAX]):
            rr = 5 + i
            for c_idx, v in enumerate(row[:MSO_NCOL], start=1):
                ws.cell(row=rr, column=c_idx).value = v

        ws.freeze_panes = "A5"
        ws.protection.sheet = True
